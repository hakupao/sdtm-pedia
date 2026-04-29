#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_xlsx_text.py — .xlsx → 全文 plain text 抽出ツール

使用方法:
  python3 docs/jp/scripts/extract_xlsx_text.py <xlsx_path> [--format raw|jsonl] [--output <path>]

入力:  .xlsx ファイルパス (第 1 引数)
出力:  jsonl (既定) または raw テキスト. --output 省略時は stdout.

出力形式:
  jsonl: 1 cell 1 行 {"sheet": ..., "row": ..., "col": ..., "coord": ..., "value": ..., "hyperlink": ...}
  raw:   === <sheet> ===\n<value>\n...

依存: openpyxl (標準依存), 標準 json / argparse / sys / os
"""

import argparse
import json
import os
import sys
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ─────────────────────────────────────────────────────────────────────────────
# 公開 API
# ─────────────────────────────────────────────────────────────────────────────

def extract_cells(xlsx_path: str) -> Iterator[dict]:
    """
    .xlsx の全シートを走査し、空でないセルの情報を辞書で順次 yield する.

    Parameters
    ----------
    xlsx_path : str
        読み込む .xlsx ファイルの絶対パスまたは相対パス.

    Yields
    ------
    dict
        以下のキーを持つ辞書:
          sheet     : シート名 (str)
          row       : 行番号 (1-indexed, int)
          col       : 列番号 (1-indexed, int)
          coord     : セル座標 ("B2" 形式, str)
          value     : セル値を str 変換したもの
          hyperlink : ハイパーリンク先 URL (なければ None)

    Notes
    -----
    - MergedCell インスタンス (左上以外の結合セル) はスキップする.
    - value が None または空文字列のセルはスキップする.
    - data_only=True で数式セルは計算済み値を返す (キャッシュがない場合は None).
    """
    wb = load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                # 結合セルの左上以外はスキップ
                if isinstance(cell, MergedCell):
                    continue
                value = cell.value
                # 空セルはスキップ
                if value is None:
                    continue
                str_value = str(value)
                if str_value.strip() == "":
                    continue
                # ハイパーリンク取得
                hyperlink = None
                if cell.hyperlink is not None:
                    hyperlink = cell.hyperlink.target
                yield {
                    "sheet":     sheet_name,
                    "row":       cell.row,
                    "col":       cell.column,
                    "coord":     cell.coordinate,
                    "value":     str_value,
                    "hyperlink": hyperlink,
                }


# ─────────────────────────────────────────────────────────────────────────────
# 出力フォーマッタ
# ─────────────────────────────────────────────────────────────────────────────

def _format_jsonl(cells: Iterator[dict], out) -> None:
    """各セルを JSON Lines 形式で出力する."""
    for cell in cells:
        out.write(json.dumps(cell, ensure_ascii=False) + "\n")


def _format_raw(cells: Iterator[dict], out) -> None:
    """シートごとにヘッダー行を挿入してセル値を出力する."""
    current_sheet = None
    for cell in cells:
        if cell["sheet"] != current_sheet:
            current_sheet = cell["sheet"]
            out.write(f"=== {current_sheet} ===\n")
        out.write(cell["value"] + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# CLI エントリポイント
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    """
    CLI エントリポイント.

    終了コード:
        0 : 正常完了
        2 : xlsx ファイル不在 / 引数エラー
    """
    parser = argparse.ArgumentParser(
        description=".xlsx から全セルのテキストを抽出する.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("xlsx_path", help="入力 .xlsx ファイルのパス")
    parser.add_argument(
        "--format",
        choices=["raw", "jsonl"],
        default="jsonl",
        help="出力形式 (既定: jsonl)",
    )
    parser.add_argument(
        "--output",
        metavar="PATH",
        default=None,
        help="出力先ファイルパス (省略時は stdout)",
    )
    args = parser.parse_args()

    # xlsx 存在チェック
    if not os.path.isfile(args.xlsx_path):
        print(f"[エラー] xlsx ファイルが見つからない: {args.xlsx_path}", file=sys.stderr)
        sys.exit(2)

    # セル抽出
    cells = extract_cells(args.xlsx_path)

    # 出力先決定
    if args.output:
        out = open(args.output, "w", encoding="utf-8")
    else:
        out = sys.stdout

    try:
        if args.format == "jsonl":
            _format_jsonl(cells, out)
        else:
            _format_raw(cells, out)
    finally:
        if args.output:
            out.close()


if __name__ == "__main__":
    main()
