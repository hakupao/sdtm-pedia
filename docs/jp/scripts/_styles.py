"""
_styles.py — docs/jp/scripts/ 共通スタイルモジュール
docs/jp/PLAN.md §3.1 に定義された配色 / フォント / セル書式 / 印刷設定の定数とヘルパ関数を集約する.

本モジュールは以下のスクリプトから import される:
  - generate_style_guide.py
  - generate_cover_template.py
  - build_xlsx.py
  - audit_terms.py
  - extract_xlsx_text.py
  - check_links.py

依存: openpyxl==3.1.5
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

# ─────────────────────────────────────────────────────────────────────────────
# 1. 配色定数 (PLAN §3.1 — HEX 9 種)
#    generate_style_guide.py の PALETTE / generate_cover_template.py の COLOR_* を統合
# ─────────────────────────────────────────────────────────────────────────────

# 6 桁 HEX (openpyxl PatternFill の fgColor に直接渡す形式)
COLORS = {
    "main_accent":  "1F4E79",  # メインアクセント (見出し帯 / 深ティール)
    "sub_accent":   "2E86AB",  # サブアクセント (副見出し / 中明度ブルー)
    "header_bg":    "D9E7F1",  # ヘッダー行背景 (淡ブルー)
    "zebra_odd":    "F4F8FB",  # ゼブラ縞 奇数行 (極淡ブルー)
    "border_line":  "A6B5C2",  # 罫線 (中明度グレー)
    "pass_green":   "4D8B7C",  # 強調 合格 (医療グリーン)
    "warn_orange":  "C97B5C",  # 警告 要対応 (落ち着いたオレンジ)
    "bg_white":     "FFFFFF",  # 背景 (純白)
    "caption_gray": "5A6B7A",  # 補助テキスト (グレー)
}

# 8 桁 ARGB (generate_cover_template.py との互換; アルファは常に FF)
COLOR_MAIN_ACCENT  = "FF" + COLORS["main_accent"]
COLOR_SUB_ACCENT   = "FF" + COLORS["sub_accent"]
COLOR_HEADER_BG    = "FF" + COLORS["header_bg"]
COLOR_ZEBRA        = "FF" + COLORS["zebra_odd"]
COLOR_BORDER       = "FF" + COLORS["border_line"]
COLOR_PASS_GREEN   = "FF" + COLORS["pass_green"]
COLOR_WARNING      = "FF" + COLORS["warn_orange"]
COLOR_BG_WHITE     = "FF" + COLORS["bg_white"]
COLOR_AUX_TEXT     = "FF" + COLORS["caption_gray"]
COLOR_TEXT_WHITE   = "FFFFFFFF"

# ─────────────────────────────────────────────────────────────────────────────
# 2. フォント定数 (PLAN §3.1)
# ─────────────────────────────────────────────────────────────────────────────

FONTS = {
    "jp_primary":  "Yu Gothic UI",   # 日本語 第一候補
    "jp_fallback": "Meiryo UI",      # 日本語 代替
    "jp_fallback2": "Noto Sans JP",  # 日本語 代替 2
    "en_primary":  "Segoe UI",       # 英数字
    "mono":        "Consolas",       # 等幅 (コード / パス)
}

# サイズ (pt)
SIZE_TITLE   = 20   # 表紙タイトル
SIZE_HEADING = 14   # 章見出し
SIZE_BODY    = 11   # 本文
SIZE_NOTE    = 9    # 注記

# 行高 (pt)
ROW_HEIGHT_BODY    = 18
ROW_HEIGHT_HEADING = 24

# ─────────────────────────────────────────────────────────────────────────────
# 3. セル書式オブジェクト (よく使う組合せをそのままインスタンス化)
# ─────────────────────────────────────────────────────────────────────────────

def _side(color: str = COLOR_BORDER, style: str = "thin") -> Side:
    return Side(border_style=style, color=color)


def THIN_BORDER(color: str = COLOR_BORDER) -> Border:
    """thin 罫線 (全辺)."""
    s = _side(color, "thin")
    return Border(left=s, right=s, top=s, bottom=s)


def MEDIUM_BORDER(color: str = COLOR_BORDER) -> Border:
    """medium 罫線 (全辺)."""
    s = _side(color, "medium")
    return Border(left=s, right=s, top=s, bottom=s)


def HEADER_FILL() -> PatternFill:
    """ヘッダー行背景 (淡ブルー #D9E7F1)."""
    return PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)


def ZEBRA_FILL() -> PatternFill:
    """ゼブラ縞 奇数行 (#F4F8FB)."""
    return PatternFill(fill_type="solid", fgColor=COLOR_ZEBRA)


def ACCENT_FILL() -> PatternFill:
    """メインアクセント塗りつぶし (#1F4E79)."""
    return PatternFill(fill_type="solid", fgColor=COLOR_MAIN_ACCENT)


def SOLID_FILL(hex6: str) -> PatternFill:
    """任意 6 桁 HEX のソリッド塗りつぶし."""
    return PatternFill(fill_type="solid", fgColor=hex6)


def WHITE_FILL() -> PatternFill:
    """白背景."""
    return PatternFill(fill_type="solid", fgColor=COLOR_BG_WHITE)

# ─────────────────────────────────────────────────────────────────────────────
# 4. 印刷設定ヘルパ (PLAN §3.1: A4 縦 / 余白 normal / ヘッダー繰返し)
# ─────────────────────────────────────────────────────────────────────────────

def apply_print_settings(ws, header_text: str = "", version_text: str = "") -> None:
    """
    全シート共通印刷設定を適用する.

    Parameters
    ----------
    ws           : openpyxl Worksheet
    header_text  : ページヘッダー左欄に表示するテキスト (通常は文書名)
    version_text : ページヘッダー中央欄に表示するテキスト (通常は版数)
    """
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = False
    ws.page_margins = PageMargins(
        left=0.75, right=0.75, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_title_rows = "1:1"
    ws.oddHeader.left.text   = header_text
    ws.oddHeader.center.text = version_text
    ws.oddHeader.right.text  = "第 &P 頁 / 全 &N 頁"
    ws.oddFooter.right.text  = "機密区分: 公開"

# ─────────────────────────────────────────────────────────────────────────────
# 5. セルヘルパ関数
# ─────────────────────────────────────────────────────────────────────────────

def make_header_row(ws, row_idx: int, headers: list, height: int = ROW_HEIGHT_HEADING) -> None:
    """
    ヘッダー行を生成する.

    Parameters
    ----------
    ws       : openpyxl Worksheet
    row_idx  : 行番号 (1-origin)
    headers  : ヘッダー文字列のリスト
    height   : 行高 (pt)
    """
    ws.row_dimensions[row_idx].height = height
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=text)
        cell.font      = Font(name=FONTS["jp_primary"], size=SIZE_BODY,
                              bold=True, color="FF000000")
        cell.fill      = HEADER_FILL()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER()


def make_freeze_top_row(ws, freeze_cell: str = "A2") -> None:
    """ウィンドウ枠固定を設定する."""
    ws.freeze_panes = freeze_cell


def make_body_cell(ws, row: int, col: int, value,
                   zebra: bool = False, wrap: bool = True,
                   bold: bool = False, italic: bool = False,
                   font_color: str = "FF000000",
                   font_size: int = SIZE_BODY,
                   font_name: str = None,
                   height: int = ROW_HEIGHT_BODY) -> None:
    """
    本文セルを書き込む.

    Parameters
    ----------
    ws         : openpyxl Worksheet
    row, col   : セル座標 (1-origin)
    value      : セル値
    zebra      : True なら奇数行ゼブラ色
    wrap       : テキスト折り返し
    bold, italic, font_color, font_size, font_name : フォント指定
    height     : 行高 (pt) — 既存行高を上書きしない場合は None を渡す
    """
    if font_name is None:
        font_name = FONTS["jp_primary"]
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = ZEBRA_FILL() if zebra else WHITE_FILL()
    cell.border    = THIN_BORDER()
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.font      = Font(name=font_name, size=font_size, color=font_color,
                          bold=bold, italic=italic)
    if height is not None:
        ws.row_dimensions[row].height = height


def make_title_cell(ws, text: str, row: int = 1, col: int = 1,
                    colspan: int = 5, font_size: int = SIZE_HEADING) -> None:
    """
    シート上部のタイトルバーを設定する (メインアクセント背景 + 白文字).

    Parameters
    ----------
    ws        : openpyxl Worksheet
    text      : タイトル文字列
    row, col  : 開始セル座標
    colspan   : マージする列数
    font_size : フォントサイズ
    """
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + colspan - 1
    )
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name=FONTS["jp_primary"], size=font_size,
                          bold=True, color=COLOR_TEXT_WHITE)
    cell.fill      = ACCENT_FILL()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = ROW_HEIGHT_HEADING
