#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_xlsx.py — docs/jp/sources/*.yml → スタイル付き .xlsx 生成スクリプト

使用方法:
  python3 docs/jp/scripts/build_xlsx.py docs/jp/sources/NN.yml

入力: yml ファイルパス (第 1 引数)
出力: yml 内の document.output に指定したパス

依存: openpyxl==3.1.5, PyYAML
"""

import os
import sys
import yaml

sys.path.insert(0, os.path.dirname(__file__))
import _styles

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# placeholder トークン (generate_cover_template.py と同一仕様)
# ─────────────────────────────────────────────────────────────────────────────
PH = {
    "DOC_NAME":     "{{DOC_NAME}}",
    "DOC_NUMBER":   "{{DOC_NUMBER}}",
    "AUTHOR":       "{{AUTHOR}}",
    "CREATED_DATE": "{{CREATED_DATE}}",
    "UPDATED_DATE": "{{UPDATED_DATE}}",
    "VERSION":      "{{VERSION}}",
}


# ─────────────────────────────────────────────────────────────────────────────
# 内部ヘルパ
# ─────────────────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(border_style="thin", color=_styles.COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _medium_border() -> Border:
    s = Side(border_style="medium", color=_styles.COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _replace_ph(text: str, doc: dict) -> str:
    """yml document フィールド値で placeholder を一括置換する."""
    return (text
            .replace(PH["DOC_NAME"],     doc.get("name", ""))
            .replace(PH["DOC_NUMBER"],   doc.get("number", ""))
            .replace(PH["AUTHOR"],       doc.get("author", ""))
            .replace(PH["CREATED_DATE"], doc.get("created", ""))
            .replace(PH["UPDATED_DATE"], doc.get("updated", ""))
            .replace(PH["VERSION"],      doc.get("version", "")))


def _char_width(ch: str) -> float:
    """文字幅を返す. CJK 文字は 2.0, その他は 1.0."""
    cp = ord(ch)
    # CJK 統合漢字 / 仮名 / 全角記号などの主要ブロック
    if (0x1100 <= cp <= 0x11FF   # ハングル字母
            or 0x2E80 <= cp <= 0x2FFF   # CJK 部首補助 / 康熙部首
            or 0x3000 <= cp <= 0x9FFF   # CJK 記号〜CJK 統合漢字
            or 0xA000 <= cp <= 0xA4CF   # 彝文字
            or 0xAC00 <= cp <= 0xD7AF   # ハングル音節
            or 0xF900 <= cp <= 0xFAFF   # CJK 互換漢字
            or 0xFE10 <= cp <= 0xFE1F   # 縦書き形
            or 0xFE30 <= cp <= 0xFE4F   # CJK 互換形
            or 0xFF00 <= cp <= 0xFF60   # 全角 ASCII / 半角カタカナ
            or 0xFFE0 <= cp <= 0xFFE6   # 全角記号
            or 0x1F300 <= cp <= 0x1F9FF  # 絵文字
            or 0x20000 <= cp <= 0x2FA1F):  # CJK 拡張
        return 2.0
    return 1.0


def _cell_display_width(value) -> float:
    """セル値の表示幅を文字幅の合計で返す."""
    if value is None:
        return 0.0
    text = str(value)
    return sum(_char_width(ch) for ch in text)


def _apply_col_widths(ws, n_cols: int, all_rows_values: list,
                      manual_widths: list = None,
                      max_cap: float = 60.0, padding: float = 2.0) -> None:
    """
    列幅を内容ベースで自動調整する.

    全行のセル値文字列長を計算し、全角文字 (CJK) は幅 2.0, 半角は 1.0 で重み付けして
    列ごとの最大幅を求め、上限 max_cap + padding を加算して列幅を設定する.
    manual_widths に値がある列はその値で上書きする.

    Parameters
    ----------
    ws             : openpyxl Worksheet
    n_cols         : 列数
    all_rows_values: 全行の値リスト (list of list) — ヘッダー行を含む
    manual_widths  : yml の column_widths キー値 (省略時は None → auto のみ)
    max_cap        : 自動計算幅の上限 (文字幅単位)
    padding        : 上限適用前に加算する余白 (文字幅単位)
    """
    col_max = [0.0] * n_cols
    for row_vals in all_rows_values:
        for ci, val in enumerate(row_vals):
            if ci >= n_cols:
                break
            w = _cell_display_width(val)
            if w > col_max[ci]:
                col_max[ci] = w

    for ci in range(n_cols):
        col_letter = get_column_letter(ci + 1)
        if manual_widths and ci < len(manual_widths) and manual_widths[ci] is not None:
            ws.column_dimensions[col_letter].width = float(manual_widths[ci])
        else:
            ws.column_dimensions[col_letter].width = min(col_max[ci] + padding, max_cap)


def _apply_print_settings(ws, doc: dict) -> None:
    """全シート共通印刷設定 (A4 縦 / 余白 normal / ヘッダー繰返し)."""
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = False
    ws.page_margins = PageMargins(
        left=0.75, right=0.75, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    ws.print_title_rows = "1:1"
    ws.oddHeader.left.text   = doc.get("name", "")
    ws.oddHeader.center.text = doc.get("version", "")
    ws.oddHeader.right.text  = "第 &P 頁 / 全 &N 頁"
    ws.oddFooter.left.text   = "作成日: " + doc.get("created", "")
    ws.oddFooter.right.text  = "機密区分: 公開"


# ─────────────────────────────────────────────────────────────────────────────
# 表紙シート
# ─────────────────────────────────────────────────────────────────────────────

def _validate_schema(data: dict) -> None:
    """
    yml 読み込み後のデータ構造を検査し、仕様違反を早期に検出する.

    必須キー:
        data["document"]        : 文書メタ情報ブロック (省略不可)
        data["document"]["output"]: 出力ファイルパス (省略不可)
        data["sheets"]          : 本文シート定義リスト (省略不可; 空リストは許容)

    任意キー:
        data["revisions"], data["approval"] — 省略時はデフォルト値を使用

    シート定義任意キー:
        column_widths : list[int/float] — 列ごとの手動上書き幅 (省略時は auto-fit)
        note_rows     : list[int]       — 0-indexed 本文行番号 (注意書きスタイルを適用する行)

    失敗時挙動:
        上記のいずれかの違反を検出した場合、ValueError を raise する.
        メッセージには違反したキー名または具体的なシート名を含む.
    """
    # 必須トップレベルキー
    if "document" not in data or data["document"] is None:
        raise ValueError("yml に 'document' が必須.")
    doc = data["document"]
    if "output" not in doc or doc["output"] is None:
        raise ValueError("yml に 'document.output' が必須.")
    if "sheets" not in data or data["sheets"] is None:
        raise ValueError("yml に 'sheets' が必須.")

    # 固定シート名との衝突検査
    RESERVED = {"1_表紙", "2_改訂履歴", "3_目次", "4_承認欄"}
    CANDIDATES = {
        "1_表紙":     "本文_表紙",
        "2_改訂履歴": "本文_改訂履歴",
        "3_目次":     "本文_目次",
        "4_承認欄":   "本文_承認欄",
    }
    sheets = data["sheets"]
    seen_names = set()
    for sheet_def in sheets:
        name = sheet_def.get("name", "")
        # 空文字 / 空白のみ検査
        if not name or not name.strip():
            raise ValueError("sheets[].name に空文字または空白のみの値は使用できない.")
        # 予約シート名衝突検査
        if name in RESERVED:
            proposal = CANDIDATES.get(name, name + "_本文")
            raise ValueError(
                f"予約シート名 '{name}' は本文シートに使用できない. 候補 → '{proposal}'."
            )
        # 重複検査
        if name in seen_names:
            raise ValueError(f"sheets[].name '{name}' が重複している.")
        seen_names.add(name)
        # note_rows 型検査 (存在する場合のみ)
        note_rows = sheet_def.get("note_rows")
        if note_rows is not None:
            if not isinstance(note_rows, list):
                raise ValueError(
                    f"sheets['{name}'].note_rows は list[int] でなければならない. "
                    f"実際の型: {type(note_rows).__name__}."
                )
            for i, v in enumerate(note_rows):
                if not isinstance(v, int):
                    raise ValueError(
                        f"sheets['{name}'].note_rows[{i}] は int でなければならない. "
                        f"実際の型: {type(v).__name__}."
                    )
        # column_widths 型検査 (存在する場合のみ)
        column_widths = sheet_def.get("column_widths")
        if column_widths is not None:
            if not isinstance(column_widths, list):
                raise ValueError(
                    f"sheets['{name}'].column_widths は list[int/float] でなければならない. "
                    f"実際の型: {type(column_widths).__name__}."
                )


def _sheet_cover(wb: Workbook, doc: dict) -> None:
    """
    表紙シート (固定名 '1_表紙') を Workbook に追加する.

    必須キー (doc):
        なし (すべて任意; 欠落時は空文字で代替)

    任意キー (doc):
        name      : 文書名 (表紙帯に大見出しとして表示)
        number    : 文書番号 (帯下部に表示)
        recipient : 受信者名 (メタ情報テーブル)
        author    : 作成者 (メタ情報テーブル)
        created   : 作成日 (メタ情報テーブル)
        updated   : 最終更新日 (メタ情報テーブル)
        version   : 版数 (メタ情報テーブル)

    失敗時挙動:
        openpyxl の例外を透過する. ValueError は raise しない.
    """
    ws = wb.create_sheet("1_表紙")

    col_widths = {"A": 4, "B": 20, "C": 30, "D": 20, "E": 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 表紙帯 (行 1-4): メインアクセント背景
    for row in range(1, 5):
        ws.row_dimensions[row].height = _styles.ROW_HEIGHT_HEADING
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = _styles.ACCENT_FILL()

    # 文書名 (B2:E3)
    ws.merge_cells("B2:E3")
    title_cell = ws["B2"]
    title_cell.value = doc.get("name", "")
    title_cell.font = Font(
        name=_styles.FONTS["jp_primary"], size=_styles.SIZE_TITLE,
        bold=True, color=_styles.COLOR_TEXT_WHITE
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 文書番号 (B4:E4)
    ws.merge_cells("B4:E4")
    docnum_cell = ws["B4"]
    docnum_cell.value = "文書番号: " + doc.get("number", "")
    docnum_cell.font = Font(
        name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
        color=_styles.COLOR_TEXT_WHITE
    )
    docnum_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[5].height = _styles.ROW_HEIGHT_BODY

    # メタ情報テーブル (行 6-11)
    meta_rows = [
        ("受信者",     doc.get("recipient", "iTMS 株式会社 御中")),
        ("作成者",     doc.get("author", "")),
        ("作成日",     doc.get("created", "")),
        ("最終更新日", doc.get("updated", "")),
        ("版数",       doc.get("version", "")),
        ("",           ""),
    ]
    for i, (label, value) in enumerate(meta_rows, start=6):
        ws.row_dimensions[i].height = _styles.ROW_HEIGHT_BODY
        label_cell = ws.cell(row=i, column=2)
        label_cell.value = label
        label_cell.font = Font(name=_styles.FONTS["jp_primary"],
                               size=_styles.SIZE_BODY, bold=True, color="FF000000")
        label_cell.fill = _styles.HEADER_FILL()
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        label_cell.border = _thin_border()

        ws.merge_cells(f"C{i}:E{i}")
        value_cell = ws.cell(row=i, column=3)
        value_cell.value = value
        value_cell.font = Font(name=_styles.FONTS["jp_primary"],
                               size=_styles.SIZE_BODY, color="FF000000")
        value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        value_cell.border = _thin_border()

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# 改訂履歴シート
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_revision(wb: Workbook, doc: dict, revisions: list) -> None:
    """
    改訂履歴シート (固定名 '2_改訂履歴') を Workbook に追加する.

    必須キー (doc):
        なし

    任意キー (doc):
        name, version, created — 印刷ヘッダー / フッターに使用

    任意キー (revisions の各要素):
        version   : 改訂版数
        date      : 改訂日
        content   : 改訂内容
        author    : 作成者
        confirmer : 確認者
        approver  : 承認者

    失敗時挙動:
        openpyxl の例外を透過する. revisions が空リストの場合はヘッダー行のみ出力する.
    """
    ws = wb.create_sheet("2_改訂履歴")

    col_widths = {"A": 8, "B": 14, "C": 42, "D": 14, "E": 14, "F": 14}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    headers = ["版", "改訂日", "改訂内容", "作成", "確認", "承認"]
    ws.row_dimensions[1].height = _styles.ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                         bold=True, color="FF000000")
        cell.fill = _styles.HEADER_FILL()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.freeze_panes = "A2"

    for row_offset, rev in enumerate(revisions):
        row = 2 + row_offset
        ws.row_dimensions[row].height = _styles.ROW_HEIGHT_BODY
        fill = _styles.ZEBRA_FILL() if row_offset % 2 == 0 else _styles.WHITE_FILL()
        values = [
            rev.get("version", ""),
            rev.get("date", ""),
            rev.get("content", ""),
            rev.get("author", ""),
            rev.get("confirmer", ""),
            rev.get("approver", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = val
            cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                             color="FF000000")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = _thin_border()

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# 目次シート (本文シートへのハイパーリンク自動生成)
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_toc(wb: Workbook, doc: dict, content_sheets: list,
               include_approval: bool) -> None:
    """
    目次シート (固定名 '3_目次') を Workbook に追加し、本文シートへのハイパーリンクを自動生成する.

    必須キー (doc):
        なし

    任意キー (doc):
        name, version, created — 印刷ヘッダー / フッターに使用

    任意キー (content_sheets の各要素):
        name            : シート名 (ハイパーリンクのターゲットとして使用)
        content_summary : 目次の説明列に表示する一行概要 (省略時は name を使用)

    引数:
        include_approval : True の場合、末尾に '4_承認欄' 行を追加する

    失敗時挙動:
        openpyxl の例外を透過する. content_sheets が空の場合は固定シート行のみ出力する.
    """
    ws = wb.create_sheet("3_目次")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28

    # タイトル行
    ws.row_dimensions[1].height = _styles.ROW_HEIGHT_HEADING
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "目次"
    title.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_HEADING,
                      bold=True, color=_styles.COLOR_TEXT_WHITE)
    title.fill = _styles.ACCENT_FILL()
    title.alignment = Alignment(horizontal="center", vertical="center")

    # 注記行
    ws.row_dimensions[2].height = _styles.ROW_HEIGHT_BODY
    ws.merge_cells("A2:C2")
    note = ws["A2"]
    note.value = "(本文書のシート構成に合わせ自動生成)"
    note.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_NOTE,
                     color=_styles.COLOR_AUX_TEXT)
    note.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[3].height = _styles.ROW_HEIGHT_BODY // 2

    # 列ヘッダー行
    ws.row_dimensions[4].height = _styles.ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(["No.", "シート名", "内容"], start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                         bold=True, color="FF000000")
        cell.fill = _styles.HEADER_FILL()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.freeze_panes = "A5"

    # 固定シート群 (表紙 / 改訂履歴 / 目次)
    fixed = [
        ("1_表紙",     "表紙・文書情報"),
        ("2_改訂履歴", "版数・改訂内容の記録"),
        ("3_目次",     "本シート"),
    ]
    # 本文シート群
    body = [(s["name"], s.get("content_summary", s["name"])) for s in content_sheets]
    # 承認欄
    approval_rows = [("4_承認欄", "作成・確認・承認の署名欄")] if include_approval else []

    all_rows = fixed + body + approval_rows

    for row_offset, (sheet_name, desc) in enumerate(all_rows):
        row = 5 + row_offset
        ws.row_dimensions[row].height = _styles.ROW_HEIGHT_BODY
        fill = _styles.ZEBRA_FILL() if row_offset % 2 == 0 else _styles.WHITE_FILL()

        cell_no = ws.cell(row=row, column=1, value=str(row_offset + 1))
        cell_no.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                            color="FF000000")
        cell_no.fill = fill
        cell_no.alignment = Alignment(horizontal="center", vertical="center")
        cell_no.border = _thin_border()

        cell_name = ws.cell(row=row, column=2, value=sheet_name)
        cell_name.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                              color=_styles.COLOR_SUB_ACCENT, underline="single")
        cell_name.fill = fill
        cell_name.alignment = Alignment(horizontal="left", vertical="center")
        cell_name.border = _thin_border()
        if sheet_name != "3_目次":
            cell_name.hyperlink = f"#{sheet_name}!A1"

        cell_desc = ws.cell(row=row, column=3, value=desc)
        cell_desc.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                              color="FF000000")
        cell_desc.fill = fill
        cell_desc.alignment = Alignment(horizontal="left", vertical="center")
        cell_desc.border = _thin_border()

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# 本文シート: type=text
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_text(wb: Workbook, doc: dict, sheet_def: dict) -> None:
    """
    本文シート (type=text) を Workbook に追加する. 1 行ずつ改行分割して縦に書き込む.

    必須キー (sheet_def):
        name    : シート名

    任意キー (sheet_def):
        content       : 本文テキスト (改行区切り; 省略時は空シートを生成)
        note_rows     : list[int] — 0-indexed 本文行番号 (ヘッダー行を除く). 該当行は
                        背景 ZEBRA_FILL + フォント色 caption_gray + 斜体で描画する.
        column_widths : list[int/float] — 列ごとの手動上書き幅 (省略時は auto-fit)

    任意キー (doc):
        name, version, created — 印刷ヘッダー / フッターに使用

    失敗時挙動:
        openpyxl の例外を透過する. content が空の場合はタイトル行のみ出力する.
    """
    ws = wb.create_sheet(sheet_def["name"])

    # タイトル行
    _styles.make_title_cell(ws, sheet_def["name"], row=1, col=1, colspan=2,
                            font_size=_styles.SIZE_HEADING)

    # 本文テキスト (行ごとに分割して書き込む)
    content = sheet_def.get("content", "")
    lines = [line for line in content.split("\n") if line.strip()]
    note_rows_set = set(sheet_def.get("note_rows", []))

    all_row_values = [["", sheet_def["name"]]]  # タイトル行
    for i, line in enumerate(lines):
        row = 2 + i
        ws.row_dimensions[row].height = _styles.ROW_HEIGHT_BODY
        val = line.strip()
        is_note = i in note_rows_set
        cell = ws.cell(row=row, column=2, value=val)
        if is_note:
            cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                             color=_styles.COLOR_AUX_TEXT, italic=True)
            cell.fill = _styles.ZEBRA_FILL()
        else:
            cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                             color="FF000000")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        all_row_values.append(["", val])

    # 列幅設定: A=左余白固定 2, B=auto-fit (手動上書き可)
    manual = sheet_def.get("column_widths")
    ws.column_dimensions["A"].width = manual[0] if manual and len(manual) > 0 else 2
    if manual and len(manual) > 1 and manual[1] is not None:
        ws.column_dimensions["B"].width = float(manual[1])
    else:
        max_w = max((_cell_display_width(r[1]) for r in all_row_values), default=0.0)
        ws.column_dimensions["B"].width = min(max_w + 2.0, 60.0)

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# 本文シート: type=table
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_table(wb: Workbook, doc: dict, sheet_def: dict) -> None:
    """
    本文シート (type=table) を Workbook に追加する. ヘッダー行 + ゼブラ縞データ行を生成する.

    必須キー (sheet_def):
        name    : シート名

    任意キー (sheet_def):
        headers       : 列ヘッダー名のリスト (省略時は空リスト → 列なし)
        rows          : データ行のリスト (各要素は値のリスト; 省略時は空リスト)
                        値が '合格' の場合は医療グリーン、'要対応' の場合は警告オレンジで着色する
        note_rows     : list[int] — 0-indexed データ行番号 (ヘッダー行を除く). 該当行は
                        背景 ZEBRA_FILL + フォント色 caption_gray + 斜体で描画する.
                        '合格' / '要対応' セルは note_rows より優先されない (通常データセルのみ注意スタイル適用).
        column_widths : list[int/float] — 列ごとの手動上書き幅 (省略時は auto-fit)

    任意キー (doc):
        name, version, created — 印刷ヘッダー / フッターに使用

    失敗時挙動:
        openpyxl の例外を透過する. headers / rows ともに空の場合はタイトル行のみ出力する.
    """
    ws = wb.create_sheet(sheet_def["name"])

    headers = sheet_def.get("headers", [])
    rows_data = sheet_def.get("rows", [])
    note_rows_set = set(sheet_def.get("note_rows", []))
    col_count = len(headers)

    # タイトル行
    _styles.make_title_cell(ws, sheet_def["name"], row=1, col=1,
                            colspan=max(col_count, 1),
                            font_size=_styles.SIZE_HEADING)

    # ヘッダー行
    _styles.make_header_row(ws, 2, headers)
    ws.freeze_panes = "A3"

    # データ行
    for row_offset, row_vals in enumerate(rows_data):
        row = 3 + row_offset
        ws.row_dimensions[row].height = _styles.ROW_HEIGHT_BODY
        zebra = (row_offset % 2 == 0)
        is_note = row_offset in note_rows_set
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if is_note:
                # 注意書き行: ZEBRA_FILL + caption_gray + 斜体 (判定値も上書き)
                cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                                 color=_styles.COLOR_AUX_TEXT, italic=True)
                cell.fill = _styles.ZEBRA_FILL()
            elif val == "合格":
                cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                                 bold=True, color=_styles.COLOR_TEXT_WHITE)
                cell.fill = _styles.SOLID_FILL(_styles.COLORS["pass_green"])
            elif val == "要対応":
                cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                                 bold=True, color=_styles.COLOR_TEXT_WHITE)
                cell.fill = _styles.SOLID_FILL(_styles.COLORS["warn_orange"])
            else:
                cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                                 color="FF000000")
                cell.fill = _styles.ZEBRA_FILL() if zebra else _styles.WHITE_FILL()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = _styles.THIN_BORDER()

    # 列幅 auto-fit (手動上書き優先)
    all_rows_values = [headers] + list(rows_data)
    manual = sheet_def.get("column_widths")
    _apply_col_widths(ws, col_count, all_rows_values, manual_widths=manual)

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# 本文シート: type=links
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_links(wb: Workbook, doc: dict, sheet_def: dict) -> None:
    """
    本文シート (type=links) を Workbook に追加する. 資料名列とリンク先列の 2 列構成.

    必須キー (sheet_def):
        name  : シート名

    任意キー (sheet_def):
        links         : 参照資料のリスト (各要素は label / target を持つ辞書; 省略時は空リスト)
                        label  : 資料名 (ハイパーリンクのテキストとして表示)
                        target : リンク先 URL またはファイルパス (省略時はリンク設定なし)
        column_widths : list[int/float] — 列ごとの手動上書き幅 (省略時は auto-fit)

    任意キー (doc):
        name, version, created — 印刷ヘッダー / フッターに使用

    失敗時挙動:
        openpyxl の例外を透過する. links が空の場合はヘッダー行のみ出力する.
    """
    ws = wb.create_sheet(sheet_def["name"])

    # タイトル行
    _styles.make_title_cell(ws, sheet_def["name"], row=1, col=1, colspan=2,
                            font_size=_styles.SIZE_HEADING)

    # ヘッダー行
    _styles.make_header_row(ws, 2, ["資料名", "参照先"])
    ws.freeze_panes = "A3"

    links = sheet_def.get("links", [])
    all_rows_values = [["資料名", "参照先"]]
    for row_offset, link in enumerate(links):
        row = 3 + row_offset
        ws.row_dimensions[row].height = _styles.ROW_HEIGHT_BODY
        zebra = (row_offset % 2 == 0)

        label = link.get("label", "")
        target = link.get("target", "")
        all_rows_values.append([label, target])

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                               color=_styles.COLOR_SUB_ACCENT, underline="single")
        label_cell.fill = _styles.ZEBRA_FILL() if zebra else _styles.WHITE_FILL()
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = _styles.THIN_BORDER()
        if target:
            label_cell.hyperlink = target

        target_cell = ws.cell(row=row, column=2, value=target)
        target_cell.font = Font(name=_styles.FONTS["mono"], size=_styles.SIZE_NOTE,
                                color=_styles.COLOR_AUX_TEXT)
        target_cell.fill = _styles.ZEBRA_FILL() if zebra else _styles.WHITE_FILL()
        target_cell.alignment = Alignment(horizontal="left", vertical="center")
        target_cell.border = _styles.THIN_BORDER()

    # 列幅 auto-fit (手動上書き優先)
    manual = sheet_def.get("column_widths")
    _apply_col_widths(ws, 2, all_rows_values, manual_widths=manual)

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# 承認欄シート
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_approval(wb: Workbook, doc: dict) -> None:
    """
    承認欄シート (固定名 '4_承認欄') を Workbook に追加する.
    役割 / 氏名 / 日付 / 押印・署名 の 4 列、作成・確認・承認の 3 行構成.

    必須キー (doc):
        なし

    任意キー (doc):
        name, version, created — 印刷ヘッダー / フッターに使用

    失敗時挙動:
        openpyxl の例外を透過する. ValueError は raise しない.
    """
    ws = wb.create_sheet("4_承認欄")

    col_widths = {"A": 16, "B": 24, "C": 16, "D": 28}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # タイトル行
    ws.row_dimensions[1].height = _styles.ROW_HEIGHT_HEADING
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "承認欄"
    title.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_HEADING,
                      bold=True, color=_styles.COLOR_TEXT_WHITE)
    title.fill = _styles.ACCENT_FILL()
    title.alignment = Alignment(horizontal="center", vertical="center")

    # ヘッダー行
    ws.row_dimensions[2].height = _styles.ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(["役割", "氏名", "日付", "押印・署名"], start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                         bold=True, color="FF000000")
        cell.fill = _styles.HEADER_FILL()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _medium_border()

    # 3 行 (作成 / 確認 / 承認)
    roles = ["作成", "確認", "承認"]
    for row_offset, role in enumerate(roles):
        row = 3 + row_offset
        ws.row_dimensions[row].height = 24
        fill = _styles.ZEBRA_FILL() if row_offset % 2 == 0 else _styles.WHITE_FILL()
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = fill
            cell.font = Font(name=_styles.FONTS["jp_primary"], size=_styles.SIZE_BODY,
                             color="FF000000")
            cell.alignment = (Alignment(horizontal="left", vertical="center", wrap_text=True)
                              if col_idx == 1
                              else Alignment(horizontal="center", vertical="center", wrap_text=True))
            cell.border = _medium_border()
        ws.cell(row=row, column=1).value = role

    _apply_print_settings(ws, doc)


# ─────────────────────────────────────────────────────────────────────────────
# メイン処理
# ─────────────────────────────────────────────────────────────────────────────

def build(yml_path: str) -> str:
    """
    yml ファイルを読み込み .xlsx を生成する.

    Parameters
    ----------
    yml_path : str
        入力 yml ファイルの絶対パスまたは相対パス (カレントディレクトリ基準)

    Returns
    -------
    str
        生成した .xlsx ファイルの絶対パス
    """
    with open(yml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    _validate_schema(data)

    doc       = data.get("document", {})
    revisions = data.get("revisions", [])
    sheets    = data.get("sheets", [])
    approval  = data.get("approval", {})
    include_approval = approval.get("enabled", False)

    # 出力パスの解決 (yml の output は repo root 相対パスとして扱う)
    output_raw = doc.get("output", "docs/jp/templates/output.xlsx")
    repo_root = os.path.normpath(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..")
    )
    output_path = os.path.normpath(os.path.join(repo_root, output_raw))
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    # デフォルトシートを削除
    if wb.active is not None and wb.active.title == "Sheet":
        del wb["Sheet"]

    # 1. 表紙
    _sheet_cover(wb, doc)

    # 2. 改訂履歴
    _sheet_revision(wb, doc, revisions)

    # 3. 目次 (本文シートリストを渡してハイパーリンク自動生成)
    _sheet_toc(wb, doc, sheets, include_approval)

    # 4. 本文シート群 (type 別に分岐)
    type_handlers = {
        "text":  _sheet_text,
        "table": _sheet_table,
        "links": _sheet_links,
    }
    for sheet_def in sheets:
        sheet_type = sheet_def.get("type", "text")
        handler = type_handlers.get(sheet_type)
        if handler is None:
            print(f"[警告] 未対応のシートタイプ: {sheet_type} (シート: {sheet_def.get('name')})",
                  file=sys.stderr)
            continue
        handler(wb, doc, sheet_def)

    # 5. 承認欄
    if include_approval:
        _sheet_approval(wb, doc)

    wb.save(output_path)
    return output_path


def main() -> None:
    """
    CLI エントリポイント. yml ファイルパスを引数に取り、build() を呼び出して xlsx を生成する.

    CLI 引数:
        第 1 引数 (必須): 入力 yml ファイルのパス

    終了コード:
        0 : 正常生成完了
        1 : 引数不足 / ファイル不在 / スキーマ違反 (ValueError) / その他例外

    失敗時挙動:
        ValueError は stderr に '[エラー]' プレフィックス付きでメッセージを出力して終了する.
        ファイル不在は '[エラー]' プレフィックス付きメッセージを出力して終了する.
    """
    if len(sys.argv) < 2:
        print("使用方法: python3 build_xlsx.py <path/to/source.yml>", file=sys.stderr)
        sys.exit(1)

    yml_path = sys.argv[1]
    if not os.path.isfile(yml_path):
        print(f"[エラー] yml ファイルが見つからない: {yml_path}", file=sys.stderr)
        sys.exit(1)

    try:
        output_path = build(yml_path)
    except ValueError as e:
        print(f"[エラー] {e}", file=sys.stderr)
        sys.exit(1)

    file_size = os.path.getsize(output_path)

    from openpyxl import load_workbook
    wb_check = load_workbook(output_path)
    sheet_names = wb_check.sheetnames

    print(f"生成完了: {output_path}")
    print(f"ファイルサイズ: {file_size:,} bytes")
    print(f"シート数: {len(sheet_names)}")
    for s in sheet_names:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
