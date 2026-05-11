"""
generate_style_guide.py — docs/jp/templates/style_guide.xlsx 生成スクリプト
依存ライブラリ: openpyxl==3.1.5
実行方法: python3 docs/jp/scripts/generate_style_guide.py
出力先:   docs/jp/templates/style_guide.xlsx
"""

# openpyxl バージョン: 3.1.5 (2024-06-28 リリース, Python 3.8 以上対応, MIT ライセンス)

import os
import sys
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import _styles  # 共通スタイルモジュール

# ---------------------------------------------------------------------------
# 配色定数 (PLAN §3.1 — HEX 値そのまま)
# _styles.COLORS から派生して本スクリプト固有の表示用メタ情報を付与する
# ---------------------------------------------------------------------------
PALETTE = {
    "main_accent":   {"hex": _styles.COLORS["main_accent"],  "name": "メインアクセント",   "usage": "見出し帯 (深ティール)"},
    "sub_accent":    {"hex": _styles.COLORS["sub_accent"],   "name": "サブアクセント",     "usage": "副見出し (中明度ブルー)"},
    "header_bg":     {"hex": _styles.COLORS["header_bg"],    "name": "ヘッダー行背景",     "usage": "ヘッダー行背景 (淡ブルー)"},
    "zebra_odd":     {"hex": _styles.COLORS["zebra_odd"],    "name": "ゼブラ縞 (奇数行)",  "usage": "ゼブラ縞背景 (極淡ブルー)"},
    "border_line":   {"hex": _styles.COLORS["border_line"],  "name": "罫線",               "usage": "thin 罫線 (中明度グレー)"},
    "pass_green":    {"hex": _styles.COLORS["pass_green"],   "name": "強調 (合格)",        "usage": "合格・適合セル (医療グリーン)"},
    "warn_orange":   {"hex": _styles.COLORS["warn_orange"],  "name": "警告 (要対応)",      "usage": "警告・対応要セル (落ち着いたオレンジ)"},
    "bg_white":      {"hex": _styles.COLORS["bg_white"],     "name": "背景",               "usage": "セル背景 (純白)"},
    "caption_gray":  {"hex": _styles.COLORS["caption_gray"], "name": "補助テキスト",       "usage": "キャプション等 (グレー)"},
}

# ---------------------------------------------------------------------------
# フォント定数 (PLAN §3.1)
# _styles.FONTS の値を参照し、本スクリプト固有の表示用リスト形式で保持する
# ---------------------------------------------------------------------------
FONTS = [
    {"name": _styles.FONTS["jp_primary"],  "lang": "日本語 第一候補",    "fallback": "Meiryo UI → Noto Sans JP"},
    {"name": _styles.FONTS["jp_fallback"], "lang": "日本語 代替",        "fallback": "Noto Sans JP"},
    {"name": _styles.FONTS["jp_fallback2"],"lang": "日本語 代替 2",      "fallback": "—"},
    {"name": _styles.FONTS["en_primary"],  "lang": "英数字",             "fallback": "Arial"},
    {"name": _styles.FONTS["mono"],        "lang": "等幅 (コード/パス)", "fallback": "Yu Gothic UI Mono"},
]
FONT_SIZES = [_styles.SIZE_TITLE, _styles.SIZE_HEADING, _styles.SIZE_BODY, _styles.SIZE_NOTE]
FONT_SIZE_LABELS = {20: "表紙タイトル (20pt)", 14: "章見出し (14pt)", 11: "本文 (11pt)", 9: "注記 (9pt)"}

# ---------------------------------------------------------------------------
# 共通ヘルパー (本スクリプト内部用 — _styles の関数をラップして旧シグネチャを維持)
# ---------------------------------------------------------------------------

def thin_border(color_hex: str = "A6B5C2") -> Border:
    s = Side(border_style="thin", color=color_hex)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(color_hex: str = "D9E7F1") -> PatternFill:
    return PatternFill("solid", fgColor=color_hex)

def solid_fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=color_hex)

def set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width

def freeze_header(ws, freeze_cell: str = "A2"):
    ws.freeze_panes = freeze_cell

def apply_print_settings(ws):
    """A4 縦 / 余白 normal / ヘッダー繰返し"""
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75,
                                  header=0.3, footer=0.3)
    ws.print_title_rows = "1:1"
    ws.oddHeader.center.text = "style_guide v1.0 / 機密区分: 公開"
    ws.oddFooter.right.text = "&P / &N"
    ws.sheet_view.zoomScale = 100

def title_row(ws, text: str, row: int = 1, col: int = 1, colspan: int = 5):
    """シート上部のタイトルセルを設定する."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Yu Gothic UI", bold=True, size=14, color="1F4E79")
    cell.fill = solid_fill("D9E7F1")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + colspan - 1
    )

def header_cell(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Yu Gothic UI", bold=True, size=11, color="1F4E79")
    cell.fill = header_fill()
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24
    return cell

def body_cell(ws, row: int, col: int, value: str, zebra: bool = False,
              wrap: bool = True, bold: bool = False, italic: bool = False,
              font_color: str = "000000", font_size: int = 11,
              font_name: str = "Yu Gothic UI"):
    cell = ws.cell(row=row, column=col, value=value)
    fill_color = "F4F8FB" if zebra else "FFFFFF"
    cell.fill = solid_fill(fill_color)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.font = Font(name=font_name, size=font_size, color=font_color,
                     bold=bold, italic=italic)
    ws.row_dimensions[row].height = 18
    return cell

# ---------------------------------------------------------------------------
# シート 0 — 目次
# ---------------------------------------------------------------------------

def sheet_toc(wb: Workbook):
    ws = wb.create_sheet("0_目次")
    apply_print_settings(ws)

    title_row(ws, "style_guide v1.0 — 目次", row=1, colspan=4)

    # 説明
    desc = ws.cell(row=2, column=1,
                   value="本ファイルは iTMS 株式会社向け納品ドキュメント群の"
                         "配色・フォント・セル書式の共通参照標準である。"
                         "Phase 1 以降の各文書作成時は本ファイルを参照すること。")
    desc.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A")
    desc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 36

    # ヘッダー行
    headers = ["シート名", "内容", "主要参照仕様", "リンク"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 3, ci, h)

    sheets_info = [
        ("1_配色パレット",  "配色 9 種のサンプル表示 (HEX / 用途 / 視認性確認)",
         "PLAN §3.1 配色パレット"),
        ("2_フォント標本",  "フォント 5 種 × サイズ 4 種のサンプル文字列",
         "PLAN §3.1 フォント"),
        ("3_セル書式",      "行高 / 列幅 / 罫線 / ゼブラ縞 / 合格・警告色 のサンプル",
         "PLAN §3.1 セル書式"),
        ("4_レイアウト見本","ヘッダー固定 / 印刷設定 / ウィンドウ枠のサンプル表",
         "PLAN §3.1 セル書式 / 印刷設定"),
        ("5_禁止例",        "使用禁止の配色・装飾パターンと注意事項",
         "PLAN §3.1 禁止配色"),
    ]

    for ri, (sname, content, ref) in enumerate(sheets_info, 4):
        zebra = (ri % 2 == 0)
        body_cell(ws, ri, 1, sname, zebra=zebra)
        body_cell(ws, ri, 2, content, zebra=zebra)
        body_cell(ws, ri, 3, ref, zebra=zebra)
        link_cell = ws.cell(row=ri, column=4, value="→ 移動")
        link_cell.hyperlink = f"#{sname}!A1"
        link_cell.font = Font(name="Yu Gothic UI", size=11, color="2E86AB", underline="single")
        link_cell.fill = solid_fill("F4F8FB" if zebra else "FFFFFF")
        link_cell.border = thin_border()
        link_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 18

    set_col_width(ws, "A", 22)
    set_col_width(ws, "B", 44)
    set_col_width(ws, "C", 28)
    set_col_width(ws, "D", 10)
    freeze_header(ws, "A4")

# ---------------------------------------------------------------------------
# シート 1 — 配色パレット
# ---------------------------------------------------------------------------

def sheet_palette(wb: Workbook):
    ws = wb.create_sheet("1_配色パレット")
    apply_print_settings(ws)

    title_row(ws, "配色パレット (PLAN §3.1)", row=1, colspan=6)

    headers = ["用途キー", "カラー名", "HEX コード", "用途説明", "実塗りつぶし見本", "黒文字視認性"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 2, ci, h)

    for ri, (key, info) in enumerate(PALETTE.items(), 3):
        zebra = (ri % 2 == 1)
        body_cell(ws, ri, 1, key, zebra=zebra, font_size=9, font_color="5A6B7A")
        body_cell(ws, ri, 2, info["name"], zebra=zebra)
        body_cell(ws, ri, 3, f"#{info['hex']}", zebra=zebra,
                  font_name="Consolas", font_size=11)
        body_cell(ws, ri, 4, info["usage"], zebra=zebra)

        # 実塗りつぶし見本セル
        swatch = ws.cell(row=ri, column=5, value="  サンプル  ")
        swatch.fill = solid_fill(info["hex"])
        swatch.border = thin_border()
        swatch.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 18

        # 黒文字視認性確認
        visibility_text = "黒文字サンプル テキスト ABC 123"
        vis_cell = ws.cell(row=ri, column=6, value=visibility_text)
        vis_cell.fill = solid_fill(info["hex"])
        vis_cell.font = Font(name="Yu Gothic UI", size=9, color="000000")
        vis_cell.border = thin_border()
        vis_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[ri].height = 18

    set_col_width(ws, "A", 18)
    set_col_width(ws, "B", 20)
    set_col_width(ws, "C", 14)
    set_col_width(ws, "D", 32)
    set_col_width(ws, "E", 16)
    set_col_width(ws, "F", 30)
    freeze_header(ws, "A3")

# ---------------------------------------------------------------------------
# シート 2 — フォント標本
# ---------------------------------------------------------------------------

def sheet_fonts(wb: Workbook):
    ws = wb.create_sheet("2_フォント標本")
    apply_print_settings(ws)

    title_row(ws, "フォント標本 (PLAN §3.1)", row=1, colspan=7)

    # 凡例
    legend = ws.cell(row=2, column=1,
                     value="【注意】 フォントが受信者環境に未インストールの場合、Excel が代替フォントで表示する。"
                           "Yu Gothic UI → Meiryo UI → Noto Sans JP の優先順で確認すること。")
    legend.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A", italic=True)
    legend.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 36

    # ヘッダー
    headers = ["フォント名", "言語用途", "代替候補", "20pt (表紙タイトル)", "14pt (章見出し)", "11pt (本文)", "9pt (注記)"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 3, ci, h)

    sample_ja = "SDTM 知識ベース — 試験データ表形式 ABC 123"
    sample_bilingual = "臨床試験データ (Clinical Trial Data) ABC 123 あいうえお"

    for ri, finfo in enumerate(FONTS, 4):
        zebra = (ri % 2 == 0)
        body_cell(ws, ri, 1, finfo["name"], zebra=zebra, font_name="Consolas", font_size=10)
        body_cell(ws, ri, 2, finfo["lang"], zebra=zebra)
        body_cell(ws, ri, 3, finfo["fallback"], zebra=zebra, font_size=9, font_color="5A6B7A")

        # サイズ別サンプル
        for ci_off, sz in enumerate(FONT_SIZES, 4):
            cell = ws.cell(row=ri, column=ci_off, value=sample_ja)
            cell.font = Font(name=finfo["name"], size=sz)
            cell.fill = solid_fill("F4F8FB" if zebra else "FFFFFF")
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = max(18, 20 if ri == 4 else 18)

    # 日英混在サンプル行
    ri_bi = 4 + len(FONTS)
    ws.row_dimensions[ri_bi].height = 24
    bi_label = ws.cell(row=ri_bi, column=1, value="日英混在確認行")
    bi_label.font = Font(name="Yu Gothic UI", bold=True, size=9, color="5A6B7A")
    bi_label.fill = solid_fill("D9E7F1")
    bi_label.border = thin_border()
    ws.merge_cells(start_row=ri_bi, start_column=1, end_row=ri_bi, end_column=3)

    for ci_off, sz in enumerate(FONT_SIZES, 4):
        cell = ws.cell(row=ri_bi, column=ci_off, value=sample_bilingual)
        cell.font = Font(name="Yu Gothic UI", size=sz)
        cell.fill = solid_fill("D9E7F1")
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    set_col_width(ws, "A", 20)
    set_col_width(ws, "B", 18)
    set_col_width(ws, "C", 20)
    set_col_width(ws, "D", 28)
    set_col_width(ws, "E", 28)
    set_col_width(ws, "F", 28)
    set_col_width(ws, "G", 28)
    freeze_header(ws, "A4")

# ---------------------------------------------------------------------------
# シート 3 — セル書式
# ---------------------------------------------------------------------------

def sheet_cells(wb: Workbook):
    ws = wb.create_sheet("3_セル書式")
    apply_print_settings(ws)

    title_row(ws, "セル書式サンプル (PLAN §3.1)", row=1, colspan=4)

    headers = ["書式種別", "サンプルセル", "設定値", "適用場面"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 2, ci, h)

    rows = [
        # (種別, 値, 設定値説明, 適用場面, 特殊スタイル関数)
        ("行高: 本文 (18pt)",  "本文行のサンプル文字列",        "row.height = 18",                  "本文全行",          None),
        ("行高: 見出し (24pt)", "見出し行のサンプル",            "row.height = 24",                  "ヘッダー行",        "heading"),
        ("折り返し表示",        "長い文章は折り返して全体を表示する設定が有効である。この行は意図的に長くしてある。",
                                "Alignment(wrap_text=True)",     "文章セル全般",      "wrap"),
        ("罫線: thin",          "thin 罫線サンプル",             "Side(border_style='thin', color='A6B5C2')",
                                "全データセル",      None),
        ("ゼブラ縞 (奇数行)",   "奇数行サンプル",                "PatternFill('solid', fgColor='F4F8FB')",
                                "一覧表の交互行",     "zebra"),
        ("ゼブラ縞 (偶数行)",   "偶数行サンプル",                "PatternFill('solid', fgColor='FFFFFF')",
                                "一覧表の交互行",     None),
        ("ヘッダー帯",          "ヘッダー行テキスト",            "PatternFill('solid', fgColor='D9E7F1') + bold",
                                "各シートのヘッダー行", "header"),
        ("合格色 (医療グリーン)", "合格 / 適合",                  "PatternFill('solid', fgColor='4D8B7C')",
                                "合格・適合セル (多用禁止)", "pass_green"),
        ("警告色 (落ち着いたオレンジ)", "要対応",               "PatternFill('solid', fgColor='C97B5C')",
                                "警告・対応要セル (多用禁止)", "warn_orange"),
        ("補助テキスト",        "キャプション・注記テキスト",    "Font(color='5A6B7A', size=9)",
                                "図表キャプション / 注記",    "caption"),
        ("等幅フォント (コード/パス)", "scripts/generate_style_guide.py",
                                "Font(name='Consolas', size=11)", "ファイルパス・コード片", "monospace"),
    ]

    for ri, (label, sample, setting, usage, style_key) in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 18

        # 種別列
        lbl = ws.cell(row=ri, column=1, value=label)
        lbl.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A")
        lbl.fill = solid_fill("F4F8FB" if ri % 2 == 1 else "FFFFFF")
        lbl.border = thin_border()
        lbl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # サンプルセル (スタイル適用)
        smp = ws.cell(row=ri, column=2, value=sample)
        smp.border = thin_border()

        if style_key == "heading":
            smp.font = Font(name="Yu Gothic UI", bold=True, size=11, color="1F4E79")
            smp.fill = solid_fill("D9E7F1")
            smp.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[ri].height = 24
        elif style_key == "wrap":
            smp.font = Font(name="Yu Gothic UI", size=11)
            smp.fill = solid_fill("FFFFFF")
            smp.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[ri].height = 40
        elif style_key == "zebra":
            smp.font = Font(name="Yu Gothic UI", size=11)
            smp.fill = solid_fill("F4F8FB")
            smp.alignment = Alignment(horizontal="left", vertical="center")
        elif style_key == "header":
            smp.font = Font(name="Yu Gothic UI", bold=True, size=11, color="1F4E79")
            smp.fill = solid_fill("D9E7F1")
            smp.alignment = Alignment(horizontal="center", vertical="center")
        elif style_key == "pass_green":
            smp.font = Font(name="Yu Gothic UI", bold=True, size=11, color="FFFFFF")
            smp.fill = solid_fill("4D8B7C")
            smp.alignment = Alignment(horizontal="center", vertical="center")
        elif style_key == "warn_orange":
            smp.font = Font(name="Yu Gothic UI", bold=True, size=11, color="FFFFFF")
            smp.fill = solid_fill("C97B5C")
            smp.alignment = Alignment(horizontal="center", vertical="center")
        elif style_key == "caption":
            smp.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A", italic=True)
            smp.fill = solid_fill("FFFFFF")
            smp.alignment = Alignment(horizontal="left", vertical="center")
        elif style_key == "monospace":
            smp.font = Font(name="Consolas", size=11, color="1F4E79")
            smp.fill = solid_fill("F4F8FB")
            smp.alignment = Alignment(horizontal="left", vertical="center")
        else:
            smp.font = Font(name="Yu Gothic UI", size=11)
            smp.fill = solid_fill("F4F8FB" if ri % 2 == 1 else "FFFFFF")
            smp.alignment = Alignment(horizontal="left", vertical="center")

        # 設定値列
        set_cell = ws.cell(row=ri, column=3, value=setting)
        set_cell.font = Font(name="Consolas", size=9, color="5A6B7A")
        set_cell.fill = solid_fill("F4F8FB" if ri % 2 == 1 else "FFFFFF")
        set_cell.border = thin_border()
        set_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 適用場面列
        use_cell = ws.cell(row=ri, column=4, value=usage)
        use_cell.font = Font(name="Yu Gothic UI", size=10)
        use_cell.fill = solid_fill("F4F8FB" if ri % 2 == 1 else "FFFFFF")
        use_cell.border = thin_border()
        use_cell.alignment = Alignment(horizontal="left", vertical="center")

    set_col_width(ws, "A", 28)
    set_col_width(ws, "B", 36)
    set_col_width(ws, "C", 42)
    set_col_width(ws, "D", 26)
    freeze_header(ws, "A3")

# ---------------------------------------------------------------------------
# シート 4 — レイアウト見本
# ---------------------------------------------------------------------------

def sheet_layout(wb: Workbook):
    ws = wb.create_sheet("4_レイアウト見本")
    apply_print_settings(ws)

    title_row(ws, "レイアウト見本 (ヘッダー固定 / 印刷設定)", row=1, colspan=6)

    # 説明帯
    desc = ws.cell(row=2, column=1,
                   value="【操作確認】 このシートは「表示 → ウィンドウ枠の固定」が行 3 で設定されている。"
                         "縦スクロール時に行 1-2 (タイトル + 列見出し) が常時表示されることを確認すること。"
                         "印刷プレビューでは A4 縦、余白 normal、1 行目繰返しが適用される。")
    desc.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A", italic=True)
    desc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 40

    # サンプルデータ表 — ヘッダー
    col_headers = ["No.", "SDTM ドメイン", "領域区分", "変数数 (参考)", "規格バージョン", "備考"]
    for ci, h in enumerate(col_headers, 1):
        header_cell(ws, 3, ci, h)

    # サンプルデータ行 (20 行: スクロール確認用)
    sample_data = [
        ("DM",  "被験者情報",     "特別目的",   24, "SDTMIG v3.4"),
        ("AE",  "有害事象",       "イベント",   28, "SDTMIG v3.4"),
        ("CM",  "併用薬",         "介入",       20, "SDTMIG v3.4"),
        ("EX",  "試験薬曝露",     "介入",       14, "SDTMIG v3.4"),
        ("LB",  "臨床検査",       "所見",       36, "SDTMIG v3.4"),
        ("MH",  "既往歴",         "イベント",   12, "SDTMIG v3.4"),
        ("VS",  "バイタルサイン", "所見",       18, "SDTMIG v3.4"),
        ("DS",  "試験中止",       "イベント",   14, "SDTMIG v3.4"),
        ("IE",  "組入除外基準",   "特別目的",   10, "SDTMIG v3.4"),
        ("SV",  "被験者来院",     "特別目的",   10, "SDTMIG v3.4"),
        ("TU",  "腫瘍同定",       "所見",       16, "SDTMIG v3.4 Oncology"),
        ("TR",  "腫瘍応答",       "所見",       20, "SDTMIG v3.4 Oncology"),
        ("RS",  "応答評価",       "所見",       14, "SDTMIG v3.4 Oncology"),
        ("EG",  "心電図",         "所見",       22, "SDTMIG v3.4"),
        ("PE",  "身体検査",       "所見",       12, "SDTMIG v3.4"),
        ("QS",  "質問票",         "所見",       10, "SDTMIG v3.4"),
        ("SC",  "被験者特性",     "特別目的",   10, "SDTMIG v3.4"),
        ("SE",  "被験者要素",     "特別目的",   10, "SDTMIG v3.4"),
        ("TA",  "試験アーム",     "試験設計",   8,  "SDTMIG v3.4"),
        ("TE",  "試験要素",       "試験設計",   8,  "SDTMIG v3.4"),
    ]

    for ri, (domain, domain_jp, category, var_count, version) in enumerate(sample_data, 4):
        zebra = (ri % 2 == 0)
        body_cell(ws, ri, 1, str(ri - 3), zebra=zebra,
                  font_name="Consolas", font_size=10)
        body_cell(ws, ri, 2, domain, zebra=zebra,
                  font_name="Consolas", font_size=11, bold=True, font_color="1F4E79")
        body_cell(ws, ri, 3, domain_jp, zebra=zebra)
        body_cell(ws, ri, 4, category, zebra=zebra)
        body_cell(ws, ri, 5, str(var_count), zebra=zebra,
                  font_name="Consolas", font_size=10)
        body_cell(ws, ri, 6, version, zebra=zebra,
                  font_size=9, font_color="5A6B7A")

    freeze_header(ws, "A4")
    set_col_width(ws, "A", 6)
    set_col_width(ws, "B", 18)
    set_col_width(ws, "C", 16)
    set_col_width(ws, "D", 14)
    set_col_width(ws, "E", 20)
    set_col_width(ws, "F", 24)

    # 印刷設定の注記
    note_row = 4 + len(sample_data) + 1
    note = ws.cell(row=note_row, column=1,
                   value="【印刷設定確認】 印刷プレビューで以下を確認すること:\n"
                         "  1. 用紙サイズ: A4 縦\n"
                         "  2. 余白: 上下 0.75 in / 左右 0.7 in (normal)\n"
                         "  3. ヘッダー: 「style_guide v1.0 / 機密区分: 公開」(中央)\n"
                         "  4. 1 行目がすべてのページで繰り返される\n"
                         "  5. ウィンドウ枠固定: 行 3 まで固定")
    note.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A", italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(
        start_row=note_row, start_column=1,
        end_row=note_row, end_column=6
    )
    ws.row_dimensions[note_row].height = 80

# ---------------------------------------------------------------------------
# シート 5 — 禁止例
# ---------------------------------------------------------------------------

def sheet_prohibited(wb: Workbook):
    ws = wb.create_sheet("5_禁止例")
    apply_print_settings(ws)

    title_row(ws, "禁止配色・装飾パターン一覧 (PLAN §3.1)", row=1, colspan=5)

    warn_note = ws.cell(row=2, column=1,
                        value="【注意】 以下の配色・装飾は本プロジェクトの納品物に使用してはならない。"
                              "医療・IT 分野の資料では信頼性と可読性を損なうため禁止とする。"
                              "代替として §3.1 配色パレット (シート「1_配色パレット」参照) を使用すること。")
    warn_note.font = Font(name="Yu Gothic UI", size=9, color="C97B5C", italic=True)
    warn_note.fill = solid_fill("FFF3ED")
    warn_note.border = thin_border("C97B5C")
    warn_note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 50

    headers = ["禁止カテゴリ", "禁止例サンプル", "理由", "代替案", "参照"]
    for ci, h in enumerate(headers, 1):
        header_cell(ws, 3, ci, h)

    prohibited_items = [
        {
            "category": "鮮やかな赤",
            "sample_value": "× 使用禁止: 赤色テキスト・セル",
            "sample_fill": "FF0000",
            "sample_font_color": "FFFFFF",
            "reason": "医療分野では赤は「緊急・危険」の意味合いが強い。過剰な警告色は読者の注意を過度に引き、"
                      "文書の信頼性を損なう。",
            "alternative": "警告色には #C97B5C (落ち着いたオレンジ) を使用すること。",
            "reference": "PLAN §3.1 禁止配色",
        },
        {
            "category": "蛍光色",
            "sample_value": "× 使用禁止: 蛍光黄色",
            "sample_fill": "FFFF00",
            "sample_font_color": "000000",
            "reason": "蛍光色は印刷物では視認性が低下し、電子表示でも目に刺激が強い。"
                      "業務文書としての品質を損なう。",
            "alternative": "強調には #D9E7F1 (淡ブルー) または #4D8B7C (医療グリーン) を使用すること。",
            "reference": "PLAN §3.1 禁止配色",
        },
        {
            "category": "グラデーション",
            "sample_value": "× 使用禁止: グラデーション塗りつぶし",
            "sample_fill": "E0A0FF",
            "sample_font_color": "000000",
            "reason": "Excel のグラデーション fill は印刷時に再現性が低く、PDF 変換や他環境での表示が不安定になる。"
                      "一貫性のある文書品質を維持できない。",
            "alternative": "単色の PatternFill ('solid') のみ使用すること。",
            "reference": "PLAN §3.1 禁止配色",
        },
        {
            "category": "影効果",
            "sample_value": "× 使用禁止: 文字・図形への影",
            "sample_fill": "CCCCCC",
            "sample_font_color": "333333",
            "reason": "影効果はビジネス文書の可読性を低下させ、印刷時に文字がにじむ原因となる。"
                      "技術・医療文書の標準から逸脱する。",
            "alternative": "強調は太字 (bold) またはヘッダー帯 (#D9E7F1) で表現すること。",
            "reference": "PLAN §3.1 禁止配色",
        },
        {
            "category": "立体効果 (3D)",
            "sample_value": "× 使用禁止: 立体グラフ・ボタン",
            "sample_fill": "B0C4DE",
            "sample_font_color": "1F4E79",
            "reason": "立体効果・3D グラフは視覚的なノイズを増加させ、データの正確な読取を妨げる。"
                      "国際的な技術文書標準 (JIS Z 8301) でも装飾過剰は非推奨である。",
            "alternative": "グラフは 2D フラット形式のみ使用すること。図表も平面デザインで統一すること。",
            "reference": "PLAN §3.1 禁止配色 / JIS Z 8301",
        },
    ]

    for ri, item in enumerate(prohibited_items, 4):
        # 禁止カテゴリ
        cat = ws.cell(row=ri, column=1, value=item["category"])
        cat.font = Font(name="Yu Gothic UI", bold=True, size=11, color="C97B5C")
        cat.fill = solid_fill("FFF3ED")
        cat.border = thin_border("C97B5C")
        cat.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[ri].height = 60

        # 禁止例サンプルセル
        smp = ws.cell(row=ri, column=2, value=item["sample_value"])
        smp.font = Font(name="Yu Gothic UI", size=11,
                        color=item["sample_font_color"], bold=True)
        smp.fill = solid_fill(item["sample_fill"])
        smp.border = thin_border("C97B5C")
        smp.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 理由
        reason = ws.cell(row=ri, column=3, value=item["reason"])
        reason.font = Font(name="Yu Gothic UI", size=10)
        reason.fill = solid_fill("FFF3ED")
        reason.border = thin_border("C97B5C")
        reason.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 代替案
        alt = ws.cell(row=ri, column=4, value=item["alternative"])
        alt.font = Font(name="Yu Gothic UI", size=10, color="4D8B7C")
        alt.fill = solid_fill("F0FAF7")
        alt.border = thin_border("4D8B7C")
        alt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 参照
        ref = ws.cell(row=ri, column=5, value=item["reference"])
        ref.font = Font(name="Yu Gothic UI", size=9, color="5A6B7A", italic=True)
        ref.fill = solid_fill("FFF3ED")
        ref.border = thin_border("C97B5C")
        ref.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    set_col_width(ws, "A", 22)
    set_col_width(ws, "B", 26)
    set_col_width(ws, "C", 40)
    set_col_width(ws, "D", 36)
    set_col_width(ws, "E", 20)
    freeze_header(ws, "A4")

# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def main():
    output_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "templates"
    )
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "style_guide.xlsx")

    wb = Workbook()
    # デフォルトで作られる Sheet を削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_toc(wb)
    sheet_palette(wb)
    sheet_fonts(wb)
    sheet_cells(wb)
    sheet_layout(wb)
    sheet_prohibited(wb)

    wb.save(output_path)
    print(f"生成完了: {output_path}")
    print(f"シート数: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
