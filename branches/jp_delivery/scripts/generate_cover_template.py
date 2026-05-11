#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_cover_template.py
openpyxl 3.1.5 ベースの共通テンプレ生成スクリプト

産物: docs/jp/templates/cover_template.xlsx
シート構成:
  1_表紙        — 文書タイトル / 文書番号 / 受信者 / 作成者 / 日付 / 版数
  2_改訂履歴    — 版 / 改訂日 / 改訂内容 / 作成 / 確認 / 承認
  3_目次        — 各シートへのハイパーリンク (ダミー 5 行)
  4_承認欄      — 役割 / 氏名 / 日付 / 押印・署名 (3 行)
  __placeholders — placeholder 一覧 (後段スクリプト参照用)

配色・フォントは docs/jp/PLAN.md §3.1 に完全準拠.

使用方法:
  python3 docs/jp/scripts/generate_cover_template.py

依存:
  openpyxl==3.1.5 (pip install openpyxl==3.1.5)
"""

import os
import sys
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
import _styles  # 共通スタイルモジュール

# ─────────────────────────────────────────────
# §3.1 共通定数 (_styles モジュールから参照)
# ─────────────────────────────────────────────

# 配色 HEX (openpyxl は ARGB 8 桁, アルファは FF で固定)
COLOR_MAIN_ACCENT   = _styles.COLOR_MAIN_ACCENT   # メインアクセント (見出し帯)
COLOR_SUB_ACCENT    = _styles.COLOR_SUB_ACCENT    # サブアクセント (副見出し / ハイパーリンク)
COLOR_HEADER_BG     = _styles.COLOR_HEADER_BG     # ヘッダー行背景 (淡ブルー)
COLOR_ZEBRA         = _styles.COLOR_ZEBRA         # ゼブラ縞 (奇数行)
COLOR_BORDER        = _styles.COLOR_BORDER        # 罫線 (中明度グレー)
COLOR_PASS_GREEN    = _styles.COLOR_PASS_GREEN    # 強調 PASS (医療グリーン)
COLOR_WARNING       = _styles.COLOR_WARNING       # 警告 (落ち着いたオレンジ)
COLOR_BG_WHITE      = _styles.COLOR_BG_WHITE      # 背景 (純白)
COLOR_AUX_TEXT      = _styles.COLOR_AUX_TEXT      # 補助テキスト (グレー)
COLOR_TEXT_WHITE    = _styles.COLOR_TEXT_WHITE    # 白文字 (濃い背景上)

# フォント
FONT_JP_PRIMARY   = _styles.FONTS["jp_primary"]   # 日本語第一候補
FONT_EN_PRIMARY   = _styles.FONTS["en_primary"]   # 英数字
FONT_MONO         = _styles.FONTS["mono"]          # 等幅 (コード / パス)

# サイズ
SIZE_TITLE        = _styles.SIZE_TITLE    # 表紙タイトル
SIZE_HEADING      = _styles.SIZE_HEADING  # 章見出し
SIZE_BODY         = _styles.SIZE_BODY     # 本文
SIZE_NOTE         = _styles.SIZE_NOTE     # 注記

# 行高 (ポイント)
ROW_HEIGHT_BODY    = _styles.ROW_HEIGHT_BODY     # 本文行
ROW_HEIGHT_HEADING = _styles.ROW_HEIGHT_HEADING  # 見出し行

# placeholder トークン (後段スクリプトが str.replace() で置換)
PH = {
    "DOC_NAME":     "{{DOC_NAME}}",
    "DOC_NUMBER":   "{{DOC_NUMBER}}",
    "AUTHOR":       "{{AUTHOR}}",
    "CREATED_DATE": "{{CREATED_DATE}}",
    "UPDATED_DATE": "{{UPDATED_DATE}}",
    "VERSION":      "{{VERSION}}",
}


# ─────────────────────────────────────────────
# ヘルパ関数
# ─────────────────────────────────────────────

def _thin_border(color: str = COLOR_BORDER) -> Border:
    """thin 罫線 (全辺)."""
    side = Side(border_style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _medium_border(color: str = COLOR_BORDER) -> Border:
    """medium 罫線 (全辺)."""
    side = Side(border_style="medium", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)


def _accent_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_MAIN_ACCENT)


def _zebra_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLOR_ZEBRA)


def _cell_font(size: int = SIZE_BODY, bold: bool = False,
               color: str = "FF000000", name: str = FONT_JP_PRIMARY) -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def _wrap_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _wrap_left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_print_settings(ws) -> None:
    """全シート共通印刷設定: A4 縦, 余白 normal, ヘッダー繰返し."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = False
    # 余白 normal (上下左右 1.91cm ≈ 0.75in / ヘッダー 0.76cm ≈ 0.3in)
    ws.page_margins = PageMargins(
        left=0.75, right=0.75, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )
    # ヘッダー行繰返し (1 行目)
    ws.print_title_rows = "1:1"
    # ページヘッダー: 左=文書名, 中央=版数, 右=ページ番号
    ws.oddHeader.left.text   = PH["DOC_NAME"]
    ws.oddHeader.center.text = PH["VERSION"]
    ws.oddHeader.right.text  = "第 &P 頁 / 全 &N 頁"
    # ページフッター: 左=作成日, 右=機密区分
    ws.oddFooter.left.text   = "作成日: " + PH["CREATED_DATE"]
    ws.oddFooter.right.text  = "機密区分: 公開"


# ─────────────────────────────────────────────
# シート生成関数
# ─────────────────────────────────────────────

def sheet_cover(wb: Workbook) -> None:
    """1_表紙 シートを生成する."""
    ws = wb.create_sheet("1_表紙")

    # 列幅設定 (A-E)
    col_widths = {"A": 4, "B": 20, "C": 30, "D": 20, "E": 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── 表紙帯 (行 1-4): メインアクセント背景 + 白文字 ──────────────────
    for row in range(1, 5):
        ws.row_dimensions[row].height = ROW_HEIGHT_HEADING
        for col in range(1, 6):  # A-E
            cell = ws.cell(row=row, column=col)
            cell.fill = _accent_fill()

    # 表紙タイトル (B2:E3 マージ)
    ws.merge_cells("B2:E3")
    title_cell = ws["B2"]
    title_cell.value = PH["DOC_NAME"]
    title_cell.font = Font(
        name=FONT_JP_PRIMARY, size=SIZE_TITLE, bold=True, color=COLOR_TEXT_WHITE
    )
    title_cell.alignment = _wrap_center()

    # 文書番号 (B4:E4)
    ws.merge_cells("B4:E4")
    docnum_cell = ws["B4"]
    docnum_cell.value = "文書番号: " + PH["DOC_NUMBER"]
    docnum_cell.font = Font(
        name=FONT_JP_PRIMARY, size=SIZE_BODY, bold=False, color=COLOR_TEXT_WHITE
    )
    docnum_cell.alignment = _wrap_center()

    # ── 空白行 ────────────────────────────────────────────────────────
    ws.row_dimensions[5].height = ROW_HEIGHT_BODY

    # ── メタ情報テーブル (行 6-11) ───────────────────────────────────
    meta_rows = [
        ("受信者",     "iTMS 株式会社 御中"),
        ("作成者",     PH["AUTHOR"]),
        ("作成日",     PH["CREATED_DATE"]),
        ("最終更新日", PH["UPDATED_DATE"]),
        ("版数",       PH["VERSION"]),
        ("",           ""),  # 余白行
    ]
    for i, (label, value) in enumerate(meta_rows, start=6):
        ws.row_dimensions[i].height = ROW_HEIGHT_BODY
        # ラベルセル (B列)
        label_cell = ws.cell(row=i, column=2)
        label_cell.value = label
        label_cell.font  = _cell_font(bold=True)
        label_cell.fill  = _header_fill()
        label_cell.alignment = _wrap_left()
        label_cell.border = _thin_border()
        # 値セル (C列, C-E マージ)
        ws.merge_cells(f"C{i}:E{i}")
        value_cell = ws.cell(row=i, column=3)
        value_cell.value = value
        value_cell.font  = _cell_font()
        value_cell.alignment = _wrap_left()
        value_cell.border = _thin_border()

    _set_print_settings(ws)


def sheet_revision(wb: Workbook) -> None:
    """2_改訂履歴 シートを生成する."""
    ws = wb.create_sheet("2_改訂履歴")

    # 列幅
    col_widths = {
        "A": 8,   # 版
        "B": 14,  # 改訂日
        "C": 42,  # 改訂内容
        "D": 14,  # 作成
        "E": 14,  # 確認
        "F": 14,  # 承認
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    headers = ["版", "改訂日", "改訂内容", "作成", "確認", "承認"]

    # ── ヘッダー行 (行 1) ────────────────────────────────────────────
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font  = Font(
            name=FONT_JP_PRIMARY, size=SIZE_BODY, bold=True, color="FF000000"
        )
        cell.fill  = _header_fill()
        cell.alignment = _wrap_center()
        cell.border = _thin_border()

    # ウィンドウ枠固定 (1 行目)
    ws.freeze_panes = "A2"

    # ── サンプル行 (行 2) ────────────────────────────────────────────
    sample_row = ["v1.0", "YYYY-MM-DD", "初版作成", "", "", ""]
    ws.row_dimensions[2].height = ROW_HEIGHT_BODY
    for col_idx, value in enumerate(sample_row, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = value
        cell.font  = _cell_font()
        cell.fill  = _zebra_fill()
        cell.alignment = _wrap_left()
        cell.border = _thin_border()

    # ── 追加空行 (行 3-7, 偶数行は白, 奇数行はゼブラ) ───────────────
    for row in range(3, 8):
        ws.row_dimensions[row].height = ROW_HEIGHT_BODY
        fill = _zebra_fill() if row % 2 == 1 else PatternFill(
            fill_type="solid", fgColor=COLOR_BG_WHITE
        )
        for col_idx in range(1, 7):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill  = fill
            cell.font  = _cell_font()
            cell.alignment = _wrap_left()
            cell.border = _thin_border()

    _set_print_settings(ws)


def sheet_toc(wb: Workbook) -> None:
    """3_目次 シートを生成する."""
    ws = wb.create_sheet("3_目次")

    # 列幅
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20

    # ── タイトル行 ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADING
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "目次"
    title.font  = Font(
        name=FONT_JP_PRIMARY, size=SIZE_HEADING, bold=True, color=COLOR_TEXT_WHITE
    )
    title.fill  = _accent_fill()
    title.alignment = _wrap_center()

    # 注記行
    ws.row_dimensions[2].height = ROW_HEIGHT_BODY
    ws.merge_cells("A2:C2")
    note = ws["A2"]
    note.value = "(本文書のシート構成に合わせ自動生成)"
    note.font  = Font(
        name=FONT_JP_PRIMARY, size=SIZE_NOTE, bold=False, color=COLOR_AUX_TEXT
    )
    note.alignment = _wrap_left()

    ws.row_dimensions[3].height = ROW_HEIGHT_BODY / 2

    # ── 列ヘッダー行 (行 4) ──────────────────────────────────────────
    ws.row_dimensions[4].height = ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(["No.", "シート名", "内容"], start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font  = Font(
            name=FONT_JP_PRIMARY, size=SIZE_BODY, bold=True, color="FF000000"
        )
        cell.fill  = _header_fill()
        cell.alignment = _wrap_center()
        cell.border = _thin_border()

    ws.freeze_panes = "A5"

    # ── ダミーリンク 5 行 (行 5-9) ──────────────────────────────────
    dummy_sheets = [
        ("1", "1_表紙",     "表紙・文書情報"),
        ("2", "2_改訂履歴", "版数・改訂内容の記録"),
        ("3", "3_目次",     "本シート"),
        ("4", "4_承認欄",   "作成・確認・承認の署名欄"),
        ("5", "(本文シート)", "各章の本文"),
    ]
    for row_offset, (num, sheet_name, desc) in enumerate(dummy_sheets):
        row = 5 + row_offset
        ws.row_dimensions[row].height = ROW_HEIGHT_BODY
        fill = _zebra_fill() if row_offset % 2 == 0 else PatternFill(
            fill_type="solid", fgColor=COLOR_BG_WHITE
        )
        # No. 列
        cell_no = ws.cell(row=row, column=1)
        cell_no.value = num
        cell_no.font  = _cell_font()
        cell_no.fill  = fill
        cell_no.alignment = _wrap_center()
        cell_no.border = _thin_border()
        # シート名列 (ハイパーリンク付き)
        cell_name = ws.cell(row=row, column=2)
        cell_name.value = sheet_name
        cell_name.font  = Font(
            name=FONT_JP_PRIMARY, size=SIZE_BODY,
            color=COLOR_SUB_ACCENT, underline="single"
        )
        cell_name.fill  = fill
        cell_name.alignment = _wrap_left()
        cell_name.border = _thin_border()
        # 内部ハイパーリンク (実シート名にリンク; ダミー行なので anchor のみ)
        if sheet_name not in ("(本文シート)",):
            cell_name.hyperlink = f"#{sheet_name}!A1"
        # 内容列
        cell_desc = ws.cell(row=row, column=3)
        cell_desc.value = desc
        cell_desc.font  = _cell_font()
        cell_desc.fill  = fill
        cell_desc.alignment = _wrap_left()
        cell_desc.border = _thin_border()

    _set_print_settings(ws)


def sheet_approval(wb: Workbook) -> None:
    """4_承認欄 シートを生成する."""
    ws = wb.create_sheet("4_承認欄")

    # 列幅 (印影スペース確保のため D 列を広めに)
    col_widths = {
        "A": 16,  # 役割
        "B": 24,  # 氏名
        "C": 16,  # 日付
        "D": 28,  # 押印・署名
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── タイトル行 ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADING
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "承認欄"
    title.font  = Font(
        name=FONT_JP_PRIMARY, size=SIZE_HEADING, bold=True, color=COLOR_TEXT_WHITE
    )
    title.fill  = _accent_fill()
    title.alignment = _wrap_center()

    # ── ヘッダー行 (行 2) ─────────────────────────────────────────────
    ws.row_dimensions[2].height = ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(["役割", "氏名", "日付", "押印・署名"], start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font  = Font(
            name=FONT_JP_PRIMARY, size=SIZE_BODY, bold=True, color="FF000000"
        )
        cell.fill  = _header_fill()
        cell.alignment = _wrap_center()
        cell.border = _medium_border()

    # ── 3 行 (作成 / 確認 / 承認) ─────────────────────────────────────
    roles = ["作成", "確認", "承認"]
    for row_offset, role in enumerate(roles):
        row = 3 + row_offset
        # 行高 24pt (§3.2 仕様)
        ws.row_dimensions[row].height = 24
        fill = _zebra_fill() if row_offset % 2 == 0 else PatternFill(
            fill_type="solid", fgColor=COLOR_BG_WHITE
        )
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill  = fill
            cell.font  = _cell_font()
            cell.alignment = _wrap_center() if col_idx != 1 else _wrap_left()
            cell.border = _medium_border()
        ws.cell(row=row, column=1).value = role

    _set_print_settings(ws)


def sheet_placeholders(wb: Workbook) -> None:
    """__placeholders シートを生成する (後段スクリプト参照用)."""
    ws = wb.create_sheet("__placeholders")

    # 列幅
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 40

    # ── タイトル行 ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADING
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "Placeholder 一覧 (後段スクリプト参照用)"
    title.font  = Font(
        name=FONT_JP_PRIMARY, size=SIZE_HEADING, bold=True, color=COLOR_TEXT_WHITE
    )
    title.fill  = _accent_fill()
    title.alignment = _wrap_center()

    # ── ヘッダー行 ────────────────────────────────────────────────────
    ws.row_dimensions[2].height = ROW_HEIGHT_HEADING
    for col_idx, header in enumerate(
        ["トークン", "対応フィールド", "注記"], start=1
    ):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font  = Font(
            name=FONT_JP_PRIMARY, size=SIZE_BODY, bold=True, color="FF000000"
        )
        cell.fill  = _header_fill()
        cell.alignment = _wrap_center()
        cell.border = _thin_border()

    ws.freeze_panes = "A3"

    # ── placeholder 一覧 ──────────────────────────────────────────────
    placeholder_data = [
        ("{{DOC_NAME}}",     "文書名",         "例: 基本設計書"),
        ("{{DOC_NUMBER}}",   "文書番号",       "例: ITMS-SDTM-01-BASIC-DESIGN-v1.0"),
        ("{{AUTHOR}}",       "作成者",         "例: 担当部門名または氏名"),
        ("{{CREATED_DATE}}", "作成日",         "形式: YYYY-MM-DD"),
        ("{{UPDATED_DATE}}", "最終更新日",     "形式: YYYY-MM-DD"),
        ("{{VERSION}}",      "版数",           "例: v1.0 / v1.1 / v2.0"),
    ]
    for row_offset, (token, field, note) in enumerate(placeholder_data):
        row = 3 + row_offset
        ws.row_dimensions[row].height = ROW_HEIGHT_BODY
        fill = _zebra_fill() if row_offset % 2 == 0 else PatternFill(
            fill_type="solid", fgColor=COLOR_BG_WHITE
        )
        for col_idx, value in enumerate([token, field, note], start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font  = _cell_font(
                name=FONT_MONO if col_idx == 1 else FONT_JP_PRIMARY
            )
            cell.fill  = fill
            cell.alignment = _wrap_left()
            cell.border = _thin_border()

    _set_print_settings(ws)


# ─────────────────────────────────────────────
# メイン処理
# ─────────────────────────────────────────────

def main() -> None:
    """cover_template.xlsx を生成して docs/jp/templates/ に保存する."""
    output_path = os.path.join(
        os.path.dirname(__file__), "..", "templates", "cover_template.xlsx"
    )
    output_path = os.path.normpath(output_path)

    wb = Workbook()
    # デフォルトシート (Sheet) を削除
    if wb.active is not None and wb.active.title == "Sheet":
        del wb["Sheet"]

    sheet_cover(wb)
    sheet_revision(wb)
    sheet_toc(wb)
    sheet_approval(wb)
    sheet_placeholders(wb)

    wb.save(output_path)
    print(f"生成完了: {output_path}")
    print(f"シート: {[s.title for s in wb]}")


if __name__ == "__main__":
    main()
