#!/usr/bin/env python3
"""Validate Phase 1 terminology outputs against the source xlsx."""

from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Dict, List

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TERM_XLSX = REPO_ROOT / "source" / "SDTM Terminology.xlsx"
DEFAULT_SPEC_XLSX = REPO_ROOT / "source" / "SDTMIG_v3.4.xlsx"
DEFAULT_MD_ROOT = REPO_ROOT / "knowledge_base" / "terminology"
DEFAULT_REPORT = REPO_ROOT / ".codex" / "reports" / "phase1-terminology-validation.md"


def load_codelists_from_xlsx(term_xlsx_path: Path) -> Dict[str, dict]:
    wb = openpyxl.load_workbook(term_xlsx_path, read_only=True)
    ws = wb["SDTM Terminology 2022-09-30"]

    codelists: Dict[str, dict] = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, cl_code, extensible, cl_name, submission_val, synonym, definition, _nci_term = row
        if cl_code is None and extensible is not None:
            current_cl = str(code).strip()
            codelists[current_cl] = {
                "name": cl_name.strip() if cl_name else "",
                "extensible": extensible.strip() if isinstance(extensible, str) else str(extensible),
                "terms": [],
            }
        elif cl_code and str(cl_code).strip() in codelists:
            parent = str(cl_code).strip()
            codelists[parent]["terms"].append(
                {
                    "code": str(code).strip() if code else "",
                    "submission_value": str(submission_val).strip() if submission_val else "",
                    "synonym": str(synonym).strip() if synonym else "",
                    "definition": str(definition).strip() if definition else "",
                }
            )

    wb.close()
    return codelists


def split_markdown_row(line: str) -> List[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]

    cells: List[str] = []
    current: List[str] = []
    escaped = False
    for char in line:
        if escaped:
            current.append(char)
            escaped = False
            continue
        if char == "\\":
            escaped = True
            continue
        if char == "|":
            cells.append("".join(current).strip())
            current = []
            continue
        current.append(char)
    cells.append("".join(current).strip())
    return cells


def parse_terminology_markdown_file(path: Path) -> Dict[str, dict]:
    codelists: Dict[str, dict] = OrderedDict()
    current_code = None
    header_re = re.compile(r"^## (.+) \((C\d+)\)$")

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.rstrip()
        header_match = header_re.match(line)
        if header_match:
            current_code = header_match.group(2)
            try:
                source_file = str(path.relative_to(REPO_ROOT))
            except ValueError:
                source_file = str(path)
            codelists[current_code] = {
                "name": header_match.group(1),
                "extensible": "",
                "terms": [],
                "source_file": source_file,
            }
            continue

        if not current_code:
            continue

        if line.startswith("Extensible:"):
            codelists[current_code]["extensible"] = line.split(":", 1)[1].strip()
            continue

        if not line.startswith("|"):
            continue

        cells = split_markdown_row(line)
        if len(cells) != 4:
            continue
        if cells[0] in {"Code", "------"}:
            continue
        codelists[current_code]["terms"].append(
            {
                "code": cells[0],
                "submission_value": cells[1],
                "synonym": cells[2],
                "definition": cells[3],
            }
        )

    return codelists


def load_markdown_codelists(md_root: Path) -> Dict[str, dict]:
    codelists: Dict[str, dict] = OrderedDict()
    for path in sorted(md_root.rglob("*.md")):
        file_codelists = parse_terminology_markdown_file(path)
        for code, payload in file_codelists.items():
            codelists[code] = payload
    return codelists


def compare_codelists(expected: Dict[str, dict], actual: Dict[str, dict]) -> dict:
    errors: List[str] = []
    warnings: List[str] = []

    missing_codes = sorted(set(expected) - set(actual))
    extra_codes = sorted(set(actual) - set(expected))
    for code in missing_codes:
        errors.append(f"{code}: missing from markdown output")
    for code in extra_codes:
        warnings.append(f"{code}: present in markdown but not in xlsx")

    for code in sorted(set(expected) & set(actual)):
        exp = expected[code]
        act = actual[code]
        if exp["name"] != act["name"]:
            errors.append(f"{code}: name mismatch: md='{act['name']}' xlsx='{exp['name']}'")
        if exp["extensible"] != act["extensible"]:
            errors.append(
                f"{code}: extensible mismatch: md='{act['extensible']}' xlsx='{exp['extensible']}'"
            )
        if len(exp["terms"]) != len(act["terms"]):
            errors.append(f"{code}: term count mismatch: md={len(act['terms'])} xlsx={len(exp['terms'])}")

        exp_terms = {term["code"]: term for term in exp["terms"]}
        act_terms = {term["code"]: term for term in act["terms"]}
        for term_code in sorted(set(exp_terms) - set(act_terms)):
            errors.append(f"{code}/{term_code}: term missing from markdown")
        for term_code in sorted(set(act_terms) - set(exp_terms)):
            warnings.append(f"{code}/{term_code}: extra term present in markdown")

        for term_code in sorted(set(exp_terms) & set(act_terms)):
            for field in ("submission_value", "synonym", "definition"):
                if exp_terms[term_code][field] != act_terms[term_code][field]:
                    errors.append(
                        f"{code}/{term_code}: {field} mismatch: "
                        f"md='{act_terms[term_code][field]}' xlsx='{exp_terms[term_code][field]}'"
                    )

    return {
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "expected_codelists": len(expected),
            "actual_codelists": len(actual),
            "missing_codelists": len(missing_codes),
            "extra_codelists": len(extra_codes),
        },
    }


def extract_spec_c_codes(spec_xlsx_path: Path) -> set:
    wb = openpyxl.load_workbook(spec_xlsx_path, read_only=True)
    ws = wb["Variables"]
    codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ct = row[7]
        if ct and str(ct).strip():
            codes.update(re.findall(r"C\d+", str(ct)))
    wb.close()
    return codes


def load_core_codelist_codes(md_root: Path) -> set:
    core_root = md_root / "core"
    return set(load_markdown_codelists(core_root).keys())


def validate_core_coverage(spec_xlsx_path: Path, md_root: Path) -> dict:
    required = extract_spec_c_codes(spec_xlsx_path)
    actual = load_core_codelist_codes(md_root)
    missing = sorted(required - actual)
    return {
        "required_codes": len(required),
        "core_codes": len(actual),
        "missing_codes": missing,
    }


def write_report(compare_result: dict, core_result: dict, report_path: Path) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Phase 1 Terminology Validation",
        "",
        "## Summary",
        "",
        f"- Expected codelists: {compare_result['summary']['expected_codelists']}",
        f"- Markdown codelists: {compare_result['summary']['actual_codelists']}",
        f"- Missing codelists: {compare_result['summary']['missing_codelists']}",
        f"- Extra codelists: {compare_result['summary']['extra_codelists']}",
        f"- Spec-referenced C-codes: {core_result['required_codes']}",
        f"- Core terminology C-codes: {core_result['core_codes']}",
        f"- Missing core references: {len(core_result['missing_codes'])}",
        "",
        "## Errors",
        "",
    ]
    if compare_result["errors"]:
        lines.extend([f"- {item}" for item in compare_result["errors"]])
    else:
        lines.append("- None")

    lines.extend(["", "## Warnings", ""])
    warnings = list(compare_result["warnings"])
    if core_result["missing_codes"]:
        warnings.extend([f"core coverage missing: {code}" for code in core_result["missing_codes"]])
    if warnings:
        lines.extend([f"- {item}" for item in warnings])
    else:
        lines.append("- None")

    lines.extend(["", "## Raw JSON", "", "```json", json.dumps({
        "compare_result": compare_result,
        "core_result": core_result,
    }, indent=2), "```", ""])
    report_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--term-xlsx", type=Path, default=DEFAULT_TERM_XLSX)
    parser.add_argument("--spec-xlsx", type=Path, default=DEFAULT_SPEC_XLSX)
    parser.add_argument("--md-root", type=Path, default=DEFAULT_MD_ROOT)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    expected = load_codelists_from_xlsx(args.term_xlsx)
    actual = load_markdown_codelists(args.md_root)
    compare_result = compare_codelists(expected, actual)
    core_result = validate_core_coverage(args.spec_xlsx, args.md_root)
    write_report(compare_result, core_result, args.report)

    print(f"Terminology validation report written to {args.report}")
    print(json.dumps({
        "errors": len(compare_result["errors"]),
        "warnings": len(compare_result["warnings"]) + len(core_result["missing_codes"]),
        "expected_codelists": compare_result["summary"]["expected_codelists"],
        "actual_codelists": compare_result["summary"]["actual_codelists"],
    }, indent=2))

    return 1 if compare_result["errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
