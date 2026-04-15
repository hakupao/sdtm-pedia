"""
Phase 1 Step 3: Validate spec.md files against SDTMIG_v3.4.xlsx

Checks:
  - Every domain in xlsx has a spec.md
  - Every variable in xlsx appears in spec.md
  - Field values match exactly (label, type, controlled terms, role, core, notes)
  - Variable order is correct
"""

import openpyxl
import os
import re

XLSX_PATH = "source/SDTMIG_v3.4.xlsx"
SPEC_DIR = "knowledge_base/domains"


def load_xlsx_data():
    """Load all variables and datasets from xlsx."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    # Datasets
    datasets = {}
    ws = wb["Datasets"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            datasets[row[2].strip()] = {
                "class": row[1].strip() if row[1] else "",
                "label": row[3].strip() if row[3] else "",
                "structure": row[4].strip() if row[4] else "",
            }

    # Variables
    variables = {}
    ws = wb["Variables"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[3]:
            continue
        domain = row[3].strip()
        if domain not in variables:
            variables[domain] = []
        variables[domain].append({
            "order": str(row[1]).strip() if row[1] else "",
            "variable_name": row[4].strip() if row[4] else "",
            "label": row[5].strip() if row[5] else "",
            "type": row[6].strip() if row[6] else "",
            "controlled_terms": row[7].strip() if row[7] else "",
            "role": row[8].strip() if row[8] else "",
            "cdisc_notes": row[9].strip() if row[9] else "",
            "core": row[10].strip() if row[10] else "",
        })

    wb.close()
    return datasets, variables


def parse_spec_md(filepath):
    """Parse a spec.md file back into structured data."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # Parse header
    header = {}
    header_match = re.search(r'^# (\S+) — (.+)$', content, re.MULTILINE)
    if header_match:
        header["domain"] = header_match.group(1)
        header["label"] = header_match.group(2)

    meta_match = re.search(r'^> Class: (.+?) \| Structure: (.+)$', content, re.MULTILINE)
    if meta_match:
        header["class"] = meta_match.group(1)
        header["structure"] = meta_match.group(2)

    # Parse variables
    variables = []
    var_blocks = re.split(r'^### ', content, flags=re.MULTILINE)[1:]  # skip before first ###

    for block in var_blocks:
        lines = block.strip().split("\n")
        var = {"variable_name": lines[0].strip()}

        for line in lines[1:]:
            line = line.strip()
            if line.startswith("- **Order:**"):
                var["order"] = line.split(":**", 1)[1].strip()
            elif line.startswith("- **Label:**"):
                var["label"] = line.split(":**", 1)[1].strip()
            elif line.startswith("- **Type:**"):
                var["type"] = line.split(":**", 1)[1].strip()
            elif line.startswith("- **Controlled Terms:**"):
                var["controlled_terms"] = line.split(":**", 1)[1].strip()
            elif line.startswith("- **Role:**"):
                var["role"] = line.split(":**", 1)[1].strip()
            elif line.startswith("- **Core:**"):
                var["core"] = line.split(":**", 1)[1].strip()
            elif line.startswith("- **CDISC Notes:**"):
                var["cdisc_notes"] = line.split(":**", 1)[1].strip()

        variables.append(var)

    return header, variables


def validate():
    """Run all validations."""
    datasets, xlsx_vars = load_xlsx_data()
    errors = []
    warnings = []
    stats = {"domains_checked": 0, "variables_checked": 0, "fields_checked": 0}

    # Check every domain has a spec.md
    for domain in sorted(xlsx_vars.keys()):
        spec_path = os.path.join(SPEC_DIR, domain, "spec.md")
        if not os.path.exists(spec_path):
            errors.append(f"MISSING: {domain}/spec.md does not exist")
            continue

        stats["domains_checked"] += 1
        header, md_vars = parse_spec_md(spec_path)

        # Check header
        if domain in datasets:
            ds = datasets[domain]
            if header.get("label") != ds["label"]:
                errors.append(f"{domain}: header label mismatch: md='{header.get('label')}' xlsx='{ds['label']}'")
            if header.get("class") != ds["class"]:
                errors.append(f"{domain}: header class mismatch: md='{header.get('class')}' xlsx='{ds['class']}'")
            if header.get("structure") != ds["structure"]:
                errors.append(f"{domain}: header structure mismatch: md='{header.get('structure')}' xlsx='{ds['structure']}'")

        # Check variable count
        xlsx_var_list = xlsx_vars[domain]
        if len(md_vars) != len(xlsx_var_list):
            errors.append(f"{domain}: variable count mismatch: md={len(md_vars)} xlsx={len(xlsx_var_list)}")

        # Check each variable
        md_var_map = {v["variable_name"]: v for v in md_vars}

        for xlsx_var in xlsx_var_list:
            vname = xlsx_var["variable_name"]
            stats["variables_checked"] += 1

            if vname not in md_var_map:
                errors.append(f"{domain}: variable {vname} missing from spec.md")
                continue

            md_var = md_var_map[vname]

            # Check each field
            fields = ["order", "label", "type", "controlled_terms", "role", "core", "cdisc_notes"]
            for field in fields:
                stats["fields_checked"] += 1
                xlsx_val = xlsx_var.get(field, "")
                md_val = md_var.get(field, "")
                if xlsx_val != md_val:
                    errors.append(f"{domain}/{vname}: {field} mismatch: md='{md_val}' xlsx='{xlsx_val}'")

        # Check variable order
        md_var_names = [v["variable_name"] for v in md_vars]
        xlsx_var_names = [v["variable_name"] for v in xlsx_var_list]
        if md_var_names != xlsx_var_names:
            warnings.append(f"{domain}: variable order differs from xlsx")

    # Check for extra spec.md files not in xlsx
    if os.path.exists(SPEC_DIR):
        for d in sorted(os.listdir(SPEC_DIR)):
            if os.path.isdir(os.path.join(SPEC_DIR, d)) and d not in xlsx_vars:
                warnings.append(f"EXTRA: {d}/spec.md exists but domain not in xlsx")

    # Report
    print("=" * 60)
    print("SPEC.MD VALIDATION REPORT")
    print("=" * 60)
    print(f"Domains checked: {stats['domains_checked']}")
    print(f"Variables checked: {stats['variables_checked']}")
    print(f"Fields checked: {stats['fields_checked']}")
    print()

    if errors:
        print(f"ERRORS ({len(errors)}):")
        for e in errors:
            print(f"  ✗ {e}")
    else:
        print("ERRORS: None")

    print()
    if warnings:
        print(f"WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  ⚠ {w}")
    else:
        print("WARNINGS: None")

    print()
    if not errors:
        print("RESULT: PASS — all spec.md files match xlsx exactly")
    else:
        print(f"RESULT: FAIL — {len(errors)} errors found")

    return len(errors) == 0


if __name__ == "__main__":
    success = validate()
    exit(0 if success else 1)
