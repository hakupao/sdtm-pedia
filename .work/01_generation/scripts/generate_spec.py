"""
Phase 1 Step 1: Generate domains/*/spec.md from SDTMIG_v3.4.xlsx

Reads Variables and Datasets sheets, produces one spec.md per domain
in the flat (per-variable block) format defined in 02_restructure_plan.md.
"""

import openpyxl
import os
import json
from collections import OrderedDict
from datetime import datetime, timezone

XLSX_PATH = "source/SDTMIG_v3.4.xlsx"
OUTPUT_DIR = "knowledge_base/domains"
PROGRESS_PATH = ".work/progress.json"


def load_datasets(wb):
    """Load domain metadata from Datasets sheet."""
    ws = wb["Datasets"]
    datasets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # Dataset Name
            domain = row[2].strip()
            datasets[domain] = {
                "class": row[1].strip() if row[1] else "",
                "label": row[3].strip() if row[3] else "",
                "structure": row[4].strip() if row[4] else "",
            }
    return datasets


def load_variables(wb):
    """Load all variables from Variables sheet, grouped by domain."""
    ws = wb["Variables"]
    domains = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[3]:  # Column D = domain abbreviation
            continue
        domain = row[3].strip()
        if domain not in domains:
            domains[domain] = []
        domains[domain].append({
            "order": row[1],
            "variable_name": row[4].strip() if row[4] else "",
            "label": row[5].strip() if row[5] else "",
            "type": row[6].strip() if row[6] else "",
            "controlled_terms": row[7].strip() if row[7] else "",
            "role": row[8].strip() if row[8] else "",
            "cdisc_notes": row[9].strip() if row[9] else "",
            "core": row[10].strip() if row[10] else "",
        })
    return domains


def generate_spec_md(domain, dataset_info, variables):
    """Generate spec.md content for a single domain."""
    lines = []
    lines.append(f"# {domain} — {dataset_info['label']}")
    lines.append("")
    lines.append(f"> Class: {dataset_info['class']} | Structure: {dataset_info['structure']}")
    lines.append("")

    for var in variables:
        lines.append(f"### {var['variable_name']}")
        lines.append(f"- **Order:** {var['order']}")
        lines.append(f"- **Label:** {var['label']}")
        lines.append(f"- **Type:** {var['type']}")
        lines.append(f"- **Controlled Terms:** {var['controlled_terms']}")
        lines.append(f"- **Role:** {var['role']}")
        lines.append(f"- **Core:** {var['core']}")
        lines.append(f"- **CDISC Notes:** {var['cdisc_notes']}")
        lines.append("")

    return "\n".join(lines)


def update_progress(completed_files):
    """Update progress.json."""
    progress = {
        "phase": "phase1_xlsx_generation",
        "completed": completed_files,
        "current": None,
        "status": "in_progress",
        "last_updated": datetime.now(timezone.utc).isoformat(),
    }
    os.makedirs(os.path.dirname(PROGRESS_PATH), exist_ok=True)
    with open(PROGRESS_PATH, "w") as f:
        json.dump(progress, f, indent=2)


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    datasets = load_datasets(wb)
    variables = load_variables(wb)
    wb.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    completed = []
    errors = []

    for domain in sorted(variables.keys()):
        if domain not in datasets:
            errors.append(f"WARNING: Domain {domain} found in Variables but not in Datasets sheet")
            continue

        domain_dir = os.path.join(OUTPUT_DIR, domain)
        os.makedirs(domain_dir, exist_ok=True)

        content = generate_spec_md(domain, datasets[domain], variables[domain])
        spec_path = os.path.join(domain_dir, "spec.md")
        with open(spec_path, "w", encoding="utf-8") as f:
            f.write(content)

        completed.append(f"{domain}/spec.md")
        print(f"  Generated {spec_path} ({len(variables[domain])} variables)")

    # Check for domains in Datasets but not in Variables
    for domain in datasets:
        if domain not in variables:
            errors.append(f"WARNING: Domain {domain} in Datasets but has no variables")

    update_progress(completed)

    print(f"\nDone: {len(completed)} spec.md files generated")
    if errors:
        print("\nWarnings:")
        for e in errors:
            print(f"  {e}")


if __name__ == "__main__":
    main()
