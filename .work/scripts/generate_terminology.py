"""
Phase 1 Step 2: Generate terminology/ from SDTM Terminology.xlsx

Splits 1005 codelists into three tiers:
  - core/       : codelists referenced by SDTMIG spec (~147 C-codes)
  - questionnaires/ : QS-related codelists
  - supplementary/  : everything else

Within core/, groups by domain association (data-driven from spec xlsx).
Files are split to stay under 100KB.
"""

import openpyxl
import os
import re
import json
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone

TERM_XLSX = "source/SDTM Terminology.xlsx"
SPEC_XLSX = "source/SDTMIG_v3.4.xlsx"
OUTPUT_DIR = "knowledge_base/terminology"
PROGRESS_PATH = ".work/progress.json"
SIZE_LIMIT = 90 * 1024  # 90KB per file


def extract_ccode_usage(spec_xlsx_path):
    """Extract C-codes referenced in spec, with domain and variable context."""
    wb = openpyxl.load_workbook(spec_xlsx_path, read_only=True)
    ws = wb["Variables"]

    ccode_domains = defaultdict(set)  # C-code -> set of domains
    all_ccodes = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        ct = row[7]
        domain = row[3]
        if ct and ct.strip() and domain:
            codes = re.findall(r'C\d+', ct.strip())
            for c in codes:
                ccode_domains[c].add(domain.strip())
                all_ccodes.add(c)
    wb.close()
    return all_ccodes, ccode_domains


def load_codelists(term_xlsx_path):
    """Load all codelists from terminology xlsx."""
    wb = openpyxl.load_workbook(term_xlsx_path, read_only=True)
    ws = wb["SDTM Terminology 2022-09-30"]

    codelists = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, cl_code, extensible, cl_name, submission_val, synonym, definition, nci_term = row

        if cl_code is None and extensible is not None:
            current_cl = str(code).strip()
            codelists[current_cl] = {
                "name": cl_name.strip() if cl_name else "",
                "extensible": extensible.strip() if isinstance(extensible, str) else str(extensible),
                "terms": [],
            }
        elif cl_code and str(cl_code).strip() in codelists:
            parent = str(cl_code).strip()
            codelists[parent]["terms"].append({
                "code": str(code).strip() if code else "",
                "submission_value": str(submission_val).strip() if submission_val else "",
                "synonym": str(synonym).strip() if synonym else "",
                "definition": str(definition).strip() if definition else "",
            })

    wb.close()
    return codelists


def is_questionnaire_codelist(cl_name):
    """Check if a codelist is questionnaire-related."""
    qs_patterns = [
        "Questionnaire", "Scale ", "Score ", "Inventory ",
        "Assessment ", "Index ", "Survey ",
        "AIMS", "BDI", "BPI", "CIBIC", "CSSRS", "EQ-5D", "GAD",
        "HAM-", "IADL", "KCCQ", "MMSE", "MoCA", "NPI", "NRS",
        "PANSS", "PHQ", "SF-36", "UPDRS", "WOMAC",
        "Clinical Classification",
    ]
    name_upper = cl_name.upper()
    for p in qs_patterns:
        if p.upper() in name_upper:
            return True
    return False


def format_codelist_md(cl_code, cl_info):
    """Format a single codelist as markdown section."""
    lines = []
    lines.append(f"## {cl_info['name']} ({cl_code})")
    lines.append("")
    lines.append(f"Extensible: {cl_info['extensible']}")
    lines.append("")

    if cl_info["terms"]:
        lines.append("| Code | CDISC Submission Value | CDISC Synonym(s) | CDISC Definition |")
        lines.append("|------|----------------------|-------------------|------------------|")
        for term in cl_info["terms"]:
            sv = term["submission_value"].replace("|", "\\|")
            syn = term["synonym"].replace("|", "\\|")
            defn = term["definition"].replace("|", "\\|")
            lines.append(f"| {term['code']} | {sv} | {syn} | {defn} |")
    else:
        lines.append("*(No terms defined)*")

    lines.append("")
    return "\n".join(lines)


def estimate_size(text):
    return len(text.encode("utf-8"))


# --- Core grouping: data-driven by domain/theme ---

# Explicit theme mapping for cross-domain codelists (used by 3+ domains)
CROSS_DOMAIN_THEMES = {
    # General response
    "C66742": "general",   # No Yes Response (41 domains)
    "C66789": "general",   # Not Done (36 domains)
    "C99079": "general",   # Epoch (44 domains)
    "C71620": "general",   # Unit (32 domains)
    "C74456": "general",   # Anatomical Location (19 domains)
    "C99073": "general",   # Laterality (17 domains)
    "C99074": "general",   # Directionality (13 domains)
    "C96777": "general",   # Medical Evaluator Identifier (12 domains)
    "C66728": "general",   # Relation to Reference Period (9 domains)
    "C78734": "general",   # Specimen Type (11 domains)
    "C78733": "general",   # Specimen Condition (8 domains)
    "C85492": "general",   # Method (19 domains)
    "C78735": "general",   # Evaluator (19 domains)
    "C78736": "general",   # Reference Range Indicator (5 domains)
    "C66726": "interventions",  # Pharmaceutical Dosage Form (7 domains)
    "C66729": "interventions",  # Route of Administration (6 domains)
    "C71113": "interventions",  # Frequency (6 domains)
    "C71148": "interventions",  # Position (6 domains)
    "C99075": "general",   # Portion/Totality (4 domains)
    "C66734": "general",   # SDTM Domain Abbreviation (3 domains)
    "C111114": "general",  # Genetic Sample Type (3 domains)
    "C158113": "general",  # QRS Method (3 domains)
    "C181175": "general",  # Test Condition Response (3 domains)
    "C66797": "general",   # Category of Inclusion/Exclusion (2 domains)
    "C177908": "general",  # Collected Summarized Value Type (2 domains)
    "C177910": "general",  # Result Scale Response (2 domains)
    "C179588": "general",  # Result Type Response (2 domains)
    "C181170": "general",  # Test Operational Objective (2 domains)
}

# Domain abbreviation -> file grouping key
DOMAIN_FILE_MAP = {
    "AE": "ae",
    "DM": "dm",
    "DS": "disposition", "DV": "disposition",
    "CM": "interventions", "EC": "interventions", "EX": "interventions",
    "AG": "interventions", "ML": "interventions", "PR": "interventions",
    "SU": "interventions",
    "VS": "vs",
    "LB": "lb",
    "EG": "eg",
    "TA": "trial_design", "TD": "trial_design", "TE": "trial_design",
    "TI": "trial_design", "TM": "trial_design", "TS": "trial_design",
    "TV": "trial_design",
    "PC": "pk", "PP": "pk",
    "MB": "microbiology", "MS": "microbiology",
    "TR": "oncology", "TU": "oncology", "RS": "oncology",
    "FA": "findings_about", "SR": "findings_about",
    "CO": "special_purpose", "SE": "special_purpose",
    "SM": "special_purpose", "SV": "special_purpose",
    "QS": "qs",
    "MI": "mi",
    "CP": "cp",
    "GF": "gf",
    "IS": "is_domain",
    "OI": "oi",
}

GROUP_TITLES = {
    "general": ("General Cross-Domain Codelists", "Codelists used across many SDTMIG domains: response options, units, anatomical locations, laterality, specimen types, methods, and other shared terminologies."),
    "ae": ("Adverse Event Codelists", "Codelists for adverse event reporting."),
    "dm": ("Demographics Codelists", "Codelists for demographics domain."),
    "disposition": ("Disposition Codelists", "Codelists for disposition and protocol deviation domains."),
    "interventions": ("Intervention Codelists", "Codelists for intervention domains: dosage form, route, frequency, position, and related."),
    "vs": ("Vital Signs Codelists", "Codelists for vital signs test codes, names, and units."),
    "lb": ("Laboratory Codelists", "Codelists for laboratory test codes, names, and related."),
    "eg": ("ECG Codelists", "Codelists for ECG test codes, names, results, methods, and leads."),
    "trial_design": ("Trial Design Codelists", "Codelists for trial design domains: summary parameters, element types."),
    "pk": ("Pharmacokinetics Codelists", "Codelists for PK parameters, units, and analytical methods."),
    "microbiology": ("Microbiology Codelists", "Codelists for microbiology specimen and susceptibility testing."),
    "oncology": ("Oncology Codelists", "Codelists for tumor/lesion identification, properties, and response assessment."),
    "findings_about": ("Findings About Codelists", "Codelists for Findings About (FA) and Skin Response (SR) domains."),
    "special_purpose": ("Special-Purpose Domain Codelists", "Codelists for special-purpose domains: CO, SE, SM, SV."),
    "qs": ("Questionnaire Domain Codelists", "Codelists specifically referenced in QS domain spec (not questionnaire instrument codelists)."),
    "mi": ("Microscopic Findings Codelists", "Codelists for microscopic findings domain."),
    "cp": ("Cell Phenotype Codelists", "Codelists for cell phenotype findings domain."),
    "gf": ("Genomic Findings Codelists", "Codelists for genomic findings domain."),
    "is_domain": ("Immunogenicity Codelists", "Codelists for immunogenicity specimen assessments domain."),
    "oi": ("Non-host Organism Codelists", "Codelists for non-host organism identifiers domain."),
    # Catch-all
    "other": ("Other Referenced Codelists", "Codelists referenced by SDTMIG spec not classified into a specific group."),
}


def assign_core_group(ccode, ccode_domains):
    """Assign a core codelist to a group file."""
    # 1. Check explicit cross-domain theme assignment
    if ccode in CROSS_DOMAIN_THEMES:
        return CROSS_DOMAIN_THEMES[ccode]

    # 2. Single-domain codelist -> map by domain
    domains = ccode_domains.get(ccode, set())
    if len(domains) == 1:
        domain = list(domains)[0]
        return DOMAIN_FILE_MAP.get(domain, "other")

    # 3. Multi-domain but not in explicit map -> "other"
    return "other"


def write_file_with_splitting(dir_path, basename, title, description, codelist_items):
    """Write codelists to file(s), splitting if over SIZE_LIMIT."""
    # codelist_items: list of (ccode, cl_info)
    if not codelist_items:
        return []

    # Sort by codelist name
    codelist_items.sort(key=lambda x: x[1]["name"])

    # Pre-render all blocks
    blocks = [(ccode, format_codelist_md(ccode, info)) for ccode, info in codelist_items]

    # Check if single file fits
    header = f"# {title}\n\n> {description}\n\n"
    total_size = estimate_size(header) + sum(estimate_size(b) for _, b in blocks)

    files_written = []

    if total_size <= SIZE_LIMIT:
        # Single file
        filepath = os.path.join(dir_path, f"{basename}.md")
        os.makedirs(dir_path, exist_ok=True)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(header)
            for _, block in blocks:
                f.write(block + "\n")
        size = os.path.getsize(filepath)
        print(f"  {filepath} ({len(blocks)} codelists, {size / 1024:.1f} KB)")
        files_written.append(filepath)
    else:
        # Split into parts
        os.makedirs(dir_path, exist_ok=True)
        part = 1
        current_blocks = []
        current_size = 0

        for ccode, block in blocks:
            block_size = estimate_size(block)
            if current_size + block_size > SIZE_LIMIT and current_blocks:
                filepath = _write_part_file(dir_path, basename, title, part, current_blocks)
                files_written.append(filepath)
                part += 1
                current_blocks = []
                current_size = 0
            current_blocks.append(block)
            current_size += block_size

        if current_blocks:
            filepath = _write_part_file(dir_path, basename, title, part, current_blocks)
            files_written.append(filepath)

    return files_written


def _write_part_file(dir_path, basename, title, part_num, blocks):
    filepath = os.path.join(dir_path, f"{basename}_part{part_num}.md")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f"# {title} (Part {part_num})\n\n")
        f.write(f"> Codelists in this file: {len(blocks)}\n\n")
        for block in blocks:
            f.write(block + "\n")
    size = os.path.getsize(filepath)
    print(f"    {filepath} ({len(blocks)} codelists, {size / 1024:.1f} KB)")
    return filepath


def main():
    print("Loading referenced C-codes from SDTMIG spec...")
    referenced, ccode_domains = extract_ccode_usage(SPEC_XLSX)
    print(f"  Found {len(referenced)} unique C-codes referenced in spec")

    print("Loading terminology codelists...")
    codelists = load_codelists(TERM_XLSX)
    print(f"  Loaded {len(codelists)} codelists, {sum(len(v['terms']) for v in codelists.values())} terms")

    # Classify into core / questionnaires / supplementary
    core = {}
    questionnaires = {}
    supplementary = {}
    for ccode, cl_info in codelists.items():
        if ccode in referenced:
            core[ccode] = cl_info
        elif is_questionnaire_codelist(cl_info["name"]):
            questionnaires[ccode] = cl_info
        else:
            supplementary[ccode] = cl_info
    print(f"  Core: {len(core)}, Questionnaires: {len(questionnaires)}, Supplementary: {len(supplementary)}")

    all_files = []

    # --- Generate core/ ---
    print("\nGenerating core/ files...")
    # Group core codelists
    core_groups = defaultdict(list)  # group_key -> [(ccode, cl_info)]
    for ccode, cl_info in core.items():
        group = assign_core_group(ccode, ccode_domains)
        core_groups[group].append((ccode, cl_info))

    for group_key in sorted(core_groups.keys()):
        items = core_groups[group_key]
        title, desc = GROUP_TITLES.get(group_key, (f"{group_key.title()} Codelists", ""))
        files = write_file_with_splitting(
            os.path.join(OUTPUT_DIR, "core"), group_key, title, desc, items
        )
        all_files.extend(files)

    # --- Generate questionnaires/ ---
    print("\nGenerating questionnaires/ files...")
    qs_items = [(c, info) for c, info in questionnaires.items()]
    files = write_file_with_splitting(
        os.path.join(OUTPUT_DIR, "questionnaires"), "questionnaires",
        "Questionnaire Codelists",
        "Codelists related to questionnaires, scales, and clinical classification instruments.",
        qs_items,
    )
    all_files.extend(files)

    # --- Generate supplementary/ ---
    print("\nGenerating supplementary/ files...")
    sup_items = [(c, info) for c, info in supplementary.items()]
    files = write_file_with_splitting(
        os.path.join(OUTPUT_DIR, "supplementary"), "supplementary",
        "Supplementary Codelists",
        "Codelists not directly referenced by SDTMIG spec and not questionnaire-related.",
        sup_items,
    )
    all_files.extend(files)

    # Update progress
    update_progress(all_files)
    print(f"\nDone: {len(all_files)} terminology files generated")


def update_progress(term_files):
    """Update progress.json to include terminology files."""
    if os.path.exists(PROGRESS_PATH):
        with open(PROGRESS_PATH) as f:
            progress = json.load(f)
    else:
        progress = {
            "phase": "phase1_xlsx_generation",
            "completed": [],
            "current": None,
            "status": "in_progress",
        }

    # Store relative paths
    rel_files = [os.path.relpath(f, "knowledge_base") for f in term_files]
    # Remove old terminology entries and add new ones
    progress["completed"] = [
        f for f in progress["completed"] if not f.startswith("terminology/")
    ] + rel_files
    progress["last_updated"] = datetime.now(timezone.utc).isoformat()

    with open(PROGRESS_PATH, "w") as f:
        json.dump(progress, f, indent=2)


if __name__ == "__main__":
    main()
